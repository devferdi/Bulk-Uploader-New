import tkinter as tk
from tkinter import messagebox, scrolledtext
import sys
import io
import threading
from tkinter import filedialog
import os
import pandas as pd
import concurrent.futures
from bs4 import BeautifulSoup
import json
import urllib.parse
import unicodedata
import os
import time
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
import requests
import re
import numbers
import base64
import mimetypes
import traceback
from openai import OpenAI

file_lock = threading.Lock()  # 🔒 Prevents simultaneous write conflicts
DEFAULT_SHOPIFY_API_VERSION = "2026-01"



def normalize_filename(filename):
    # Normalize filename to handle special characters
    return unicodedata.normalize('NFC', filename)

def file_exists_in_folder(folder, filename):
    # Normalize filename to ensure correct access to files with umlauts or other special characters
    normalized_filename = normalize_filename(filename)
    file_path = os.path.join(folder, normalized_filename)
    return os.path.exists(file_path)


# URL-encode the filename to handle special characters like umlauts
def encode_filename(filename):
    return urllib.parse.quote(filename)


def build_file_lookup_keys(name):
    if not name:
        return set()
    base_name = os.path.basename(name)
    normalized_name = normalize_filename(base_name)
    keys = {
        base_name,
        normalized_name,
        base_name.lower(),
        normalized_name.lower(),
    }
    return {key for key in keys if key}


def is_valid_gid(gid):
    return isinstance(gid, str) and gid.startswith('gid://')


def remember_file_reference(files_dict, filename, gid, url):
    if not is_valid_gid(gid):
        return
    for key in build_file_lookup_keys(filename):
        files_dict[key] = (gid, url)


def fetch_file_reference(files_dict, filename):
    if not filename:
        return None
    for key in build_file_lookup_keys(filename):
        if key in files_dict:
            return files_dict[key]
    return None


def extract_filename_from_value(value):
    if not value:
        return ""
    parsed = urllib.parse.urlparse(str(value))
    if parsed.scheme and parsed.path:
        return os.path.basename(parsed.path)
    return os.path.basename(str(value))


def guess_mime_type(filename):
    mime_type, _ = mimetypes.guess_type(filename)
    if mime_type:
        return mime_type
    extension_map = {
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.png': 'image/png',
        '.gif': 'image/gif',
        '.webp': 'image/webp',
        '.svg': 'image/svg+xml',
        '.bmp': 'image/bmp',
        '.tiff': 'image/tiff',
        '.tif': 'image/tiff',
        '.pdf': 'application/pdf',
    }
    extension = os.path.splitext(filename)[1].lower()
    return extension_map.get(extension, 'application/octet-stream')


def resolve_asset_from_directories(filename, directories):
    if not filename:
        return None

    if os.path.isabs(filename) and os.path.exists(filename):
        return filename

    normalized_filename = normalize_filename(filename)

    for directory in directories:
        potential_paths = [
            os.path.join(directory, filename),
            os.path.join(directory, normalized_filename),
        ]
        for path in potential_paths:
            if os.path.exists(path):
                return path
    return None


def format_metafield_text_value(value):
    """Format metafield text values to avoid unintended decimal suffixes."""
    if isinstance(value, str):
        return value

    if isinstance(value, bool):
        return str(value)

    if isinstance(value, numbers.Number):
        try:
            numeric_value = float(value)
        except (TypeError, ValueError):
            return str(value)

        if pd.isna(numeric_value):
            return ""

        if float(numeric_value).is_integer():
            return str(int(numeric_value))

        formatted = ("{0:f}".format(numeric_value)).rstrip("0").rstrip(".")
        return formatted if formatted else "0"

    return str(value)


def set_dataframe_cell(df, row_index, column, value):
    try:
        df.at[row_index, column] = value
    except (TypeError, ValueError, pd.errors.LossySetitemError):
        df[column] = df[column].astype(object)
        df.at[row_index, column] = value

# Function to convert HTML to Shopify JSON
def html_to_shopify_json(html_input):
    # Parse the HTML input using BeautifulSoup
    soup = BeautifulSoup(html_input, 'html.parser')
    
    json_structure = {"type": "root", "children": []}

    def parse_element(element):
        # Handle paragraphs
        if element.name == 'p' and element.get_text(strip=True):  # Only include if paragraph has text
            paragraph = {"type": "paragraph", "children": []}
            for child in element.children:
                if isinstance(child, str):  # Plain text
                    paragraph["children"].append({"type": "text", "value": child})
                elif child.name == 'strong':  # Bold text
                    paragraph["children"].append({"type": "text", "value": child.get_text(), "bold": True})
            print(f"Parsed paragraph: {paragraph}")  # Debug print
            return paragraph

        # Handle unordered lists
        elif element.name == 'ul':
            list_items = []
            for li in element.find_all('li'):
                list_items.append({
                    "type": "list-item",
                    "children": [{"type": "text", "value": li.get_text()}]
                })
            list_element = {"type": "list", "listType": "unordered", "children": list_items}
            print(f"Parsed unordered list: {list_element}")  # Debug print
            return list_element

    # Parse each top-level element
    for element in soup.children:
        if isinstance(element, str):  # Ignore plain text nodes
            continue
        parsed_element = parse_element(element)
        if parsed_element:
            json_structure["children"].append(parsed_element)

    # Check if json_structure contains meaningful content
    if not json_structure["children"]:
        json_structure["children"].append({
            "type": "paragraph",
            "children": [{"type": "text", "value": ""}]  # Default empty content
        })
    
    print(f"Final JSON structure: {json_structure}")  # Debug print to show final output
    return json_structure



# Function to redirect output to the GUI text box
class RedirectOutput(io.StringIO):
    def __init__(self, text_area, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.text_area = text_area

        # Ensure the redirected output uses a readable, high-contrast style. Some
        # macOS Tk builds inherit a near-white foreground, which can make the
        # text effectively invisible against the default background when the app
        # is bundled. Explicit styling guarantees the log remains legible.
        self.text_area.configure(
            bg="#FAFAFA", fg="#0F0F0F", insertbackground="#0F0F0F",
            selectbackground="#0F62FE", selectforeground="#FFFFFF"
        )
        self.text_area.tag_configure("stdout", foreground="#0F0F0F")

    def write(self, msg):
        if not msg:
            return

        def append():
            if not self.text_area.winfo_exists():
                return
            self.text_area.insert(tk.END, msg, "stdout")
            self.text_area.see(tk.END)  # Auto scroll to the end

        # Tkinter isn't thread-safe. Schedule GUI updates on the main thread.
        try:
            self.text_area.after(0, append)
        except RuntimeError:
            # If the widget is being destroyed, silently ignore further writes.
            pass

    def flush(self):
        pass  # The flush method is required for compatibility with `sys.stdout`

# Function to read store credentials from a text file
def read_credentials(file_path):
    credentials = {}
    with open(file_path, 'r') as file:
        for line in file:
            line = line.strip()
            if not line or line.startswith('#') or '=' not in line:
                continue
            key, value = line.split('=', 1)
            credentials[key.strip()] = value.strip()
    return credentials


def normalize_shop_domain(shop_name):
    shop_name = (shop_name or "").strip().lower()
    if not shop_name:
        raise RuntimeError("Missing 'store_name' in credentials.txt.")

    if "://" in shop_name:
        parsed = urllib.parse.urlparse(shop_name)
        shop_name = parsed.netloc or parsed.path

    shop_name = shop_name.split("/", 1)[0].strip(".")
    if not shop_name:
        raise RuntimeError("Invalid 'store_name' in credentials.txt.")

    if shop_name.endswith(".myshopify.com"):
        shop_domain = shop_name
    else:
        shop_domain = f"{shop_name}.myshopify.com"

    allowed = set("abcdefghijklmnopqrstuvwxyz0123456789-.")
    if any(char not in allowed for char in shop_domain):
        raise RuntimeError("Invalid Shopify store domain in credentials.txt.")

    if not shop_domain.endswith(".myshopify.com"):
        raise RuntimeError("Shopify store must end with '.myshopify.com'.")

    return shop_domain


def get_shop_name(shop_name):
    shop_domain = normalize_shop_domain(shop_name)
    return shop_domain[: -len(".myshopify.com")]


def get_shopify_api_version(credentials):
    return (
        credentials.get("shopify_api_version")
        or credentials.get("api_version")
        or DEFAULT_SHOPIFY_API_VERSION
    )


def build_shopify_admin_urls(shop_domain, api_version):
    base_url = f"https://{shop_domain}/admin/api/{api_version}"
    return base_url, f"{base_url}/graphql.json"


def fetch_granted_access_scopes(shop_domain, access_token):
    if not shop_domain or not access_token:
        return None

    try:
        response = requests.get(
            f"https://{shop_domain}/admin/oauth/access_scopes.json",
            headers={"X-Shopify-Access-Token": access_token},
            timeout=20,
        )
        response.raise_for_status()
    except requests.RequestException as exc:
        print(f"⚠️ Could not fetch granted Shopify access scopes: {exc}")
        return None

    payload = response.json()
    access_scopes = payload.get("access_scopes", [])
    return [scope.get("handle") for scope in access_scopes if scope.get("handle")]


def log_shopify_access_scope_diagnostics(shopify_context, required_scopes=None):
    auth_method = shopify_context.get("auth_method") or "unknown"
    auth_errors = shopify_context.get("auth_errors") or []
    shop_domain = shopify_context.get("shop_domain") or "(unknown store)"
    granted_access_scopes = shopify_context.get("granted_access_scopes")

    print(f"Shopify auth method: {auth_method}")

    if auth_errors:
        print("⚠️ Other configured Shopify auth methods failed before fallback:")
        for auth_error in auth_errors:
            print(f"   - {auth_error}")

    if granted_access_scopes is None:
        print(f"⚠️ Could not verify granted Shopify scopes for {shop_domain}.")
        return

    unique_scopes = sorted({scope for scope in granted_access_scopes if scope})
    print(f"Granted Shopify scopes ({len(unique_scopes)}): {', '.join(unique_scopes)}")

    required_scopes = [scope for scope in (required_scopes or []) if scope]
    if not required_scopes:
        return

    missing_scopes = [scope for scope in required_scopes if scope not in unique_scopes]
    if not missing_scopes:
        print("✅ Current token includes the scopes needed for metaobject title resolution.")
        return

    print(
        "❌ Current token is missing required Shopify scopes: "
        + ", ".join(missing_scopes)
    )

    if auth_method == "Dev Dashboard client credentials":
        print(
            "➡️ Release a new app version with the updated scopes, update or reinstall "
            "the app on the store, then restart this tool so it gets a fresh token."
        )
    elif auth_method == "OAuth backend":
        print(
            "➡️ Reinstall or re-authorize the Shopify app for this store so the backend "
            "stores a fresh token with the new scopes."
        )
    elif auth_method == "credentials.txt access_token":
        print(
            "➡️ Replace the manual access token in credentials.txt with one created "
            "after the new scopes were granted."
        )


def fetch_access_token_from_backend(shop_domain, credentials):
    backend_base_url = (
        credentials.get("oauth_backend_url")
        or credentials.get("shopify_oauth_backend_url")
        or ""
    ).rstrip("/")
    agency_api_key = (
        credentials.get("agency_api_key")
        or credentials.get("backend_api_key")
        or ""
    ).strip()

    if not backend_base_url or not agency_api_key:
        return None

    token_url = (
        f"{backend_base_url}/api/shops/"
        f"{urllib.parse.quote(shop_domain, safe='')}/token"
    )

    try:
        response = requests.get(
            token_url,
            headers={"Authorization": f"Bearer {agency_api_key}"},
            timeout=20,
        )
    except requests.RequestException as exc:
        raise RuntimeError(
            f"Could not reach the OAuth backend at '{backend_base_url}': {exc}"
        ) from exc

    if response.status_code == 404:
        raise RuntimeError(
            "The OAuth backend returned 404 for the shop token endpoint. "
            f"Checked: {token_url}. This usually means either the store is not "
            "connected in the backend yet, or the deployed backend does not have "
            "the `/api/shops/{shop}/token` route."
        )

    if response.status_code in (401, 403):
        raise RuntimeError(
            "The OAuth backend rejected the agency API key. Check "
            "`agency_api_key` in `credentials.txt` and the backend configuration."
        )

    try:
        response.raise_for_status()
    except requests.HTTPError as exc:
        response_text = response.text[:500] if response is not None else str(exc)
        raise RuntimeError(
            f"OAuth backend token request failed with HTTP {response.status_code}. "
            f"Details: {response_text}"
        ) from exc

    payload = response.json()
    access_token = payload.get("access_token")
    if not access_token:
        raise RuntimeError(
            f"OAuth backend did not return an access token for store '{shop_domain}'."
        )
    return access_token


def fetch_access_token_from_dev_dashboard(shop_domain, credentials):
    client_id = (
        credentials.get("shopify_client_id")
        or credentials.get("client_id")
        or ""
    ).strip()
    client_secret = (
        credentials.get("shopify_client_secret")
        or credentials.get("client_secret")
        or ""
    ).strip()

    if not client_id or not client_secret:
        return None

    try:
        response = requests.post(
            f"https://{shop_domain}/admin/oauth/access_token",
            headers={
                "Content-Type": "application/x-www-form-urlencoded",
                "Accept": "application/json",
            },
            data={
                "grant_type": "client_credentials",
                "client_id": client_id,
                "client_secret": client_secret,
            },
            timeout=30,
        )
        response.raise_for_status()
    except requests.HTTPError as exc:
        response = exc.response
        response_text = response.text[:500] if response is not None else str(exc)
        if response is not None and response.status_code == 400:
            try:
                payload = response.json()
            except ValueError:
                payload = {}
            if payload.get("error") == "shop_not_permitted":
                raise RuntimeError(
                    "This store does not allow client-credentials for this app. "
                    "Configure 'oauth_backend_url=' and 'agency_api_key=' in "
                    "credentials.txt and use the OAuth backend flow instead."
                ) from exc
        raise RuntimeError(
            "Shopify rejected the Dev Dashboard token request. Confirm the app is "
            "installed on the store, that a version with the required scopes has "
            f"been released, and that the client ID/secret are correct. Details: {response_text}"
        ) from exc
    except requests.RequestException as exc:
        raise RuntimeError(
            f"Could not reach Shopify to request a token for '{shop_domain}': {exc}"
        ) from exc

    payload = response.json()
    access_token = payload.get("access_token")
    if not access_token:
        raise RuntimeError(
            f"Shopify did not return an access token for '{shop_domain}'."
        )
    return access_token


def load_shopify_context(credentials_path):
    credentials = read_credentials(credentials_path)

    shop_domain = normalize_shop_domain(credentials.get("store_name", ""))
    shop_name = get_shop_name(shop_domain)

    auth_errors = []
    auth_method = None
    access_token = credentials.get("access_token", "").strip()
    if access_token:
        auth_method = "credentials.txt access_token"
    if not access_token:
        try:
            access_token = fetch_access_token_from_backend(shop_domain, credentials)
            auth_method = "OAuth backend"
        except RuntimeError as exc:
            auth_errors.append(str(exc))
    if not access_token:
        try:
            access_token = fetch_access_token_from_dev_dashboard(shop_domain, credentials)
            auth_method = "Dev Dashboard client credentials"
        except RuntimeError as exc:
            auth_errors.append(str(exc))

    if not access_token:
        error_details = ""
        if auth_errors:
            error_details = "\n\nTried these auth methods:\n- " + "\n- ".join(auth_errors)
        raise RuntimeError(
            "No Shopify access token found. Configure one of these in "
            "credentials.txt: 'access_token=', or 'oauth_backend_url=' with "
            "'agency_api_key=', or 'shopify_client_id=' with "
            f"'shopify_client_secret='.{error_details}"
        )

    api_version = get_shopify_api_version(credentials)
    base_url, graphql_url = build_shopify_admin_urls(shop_domain, api_version)
    granted_access_scopes = fetch_granted_access_scopes(shop_domain, access_token)

    return {
        "credentials": credentials,
        "shop_domain": shop_domain,
        "shop_name": shop_name,
        "access_token": access_token,
        "auth_method": auth_method or "unknown",
        "auth_errors": auth_errors,
        "granted_access_scopes": granted_access_scopes,
        "api_version": api_version,
        "base_url": base_url,
        "graphql_url": graphql_url,
    }



def run_downloader_logic():
    # Get the directory where the executable or script is located
    if getattr(sys, 'frozen', False):  # If running as an EXE
        script_dir = os.path.dirname(sys.executable)
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))

    # Build the full path to 'credentials.txt'
    credentials_path = os.path.join(script_dir, 'credentials.txt')

    shopify_context = load_shopify_context(credentials_path)
    SHOP_NAME = shopify_context["shop_name"]
    ACCESS_TOKEN = shopify_context["access_token"]
    granted_access_scopes = set(shopify_context.get("granted_access_scopes") or [])
    log_shopify_access_scope_diagnostics(shopify_context)

    # Shopify API URL
    BASE_URL = shopify_context["base_url"]
    GRAPHQL_URL = shopify_context["graphql_url"]

    # Headers for API authentication
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": ACCESS_TOKEN
    }

    metaobject_label_cache = {}
    warned_missing_metaobject_scope = False

    def parse_metaobject_reference_values(value):
        if isinstance(value, list):
            return [str(item).strip() for item in value if str(item).strip()]

        try:
            if pd.isna(value):
                return []
        except TypeError:
            pass

        raw_value = str(value).strip()
        if not raw_value:
            return []

        if raw_value.startswith("["):
            try:
                parsed = json.loads(raw_value)
            except json.JSONDecodeError:
                parsed = None
            if isinstance(parsed, list):
                return [str(item).strip() for item in parsed if str(item).strip()]

        return [raw_value]

    def get_metaobject_display_label(metaobject_gid):
        nonlocal warned_missing_metaobject_scope

        if not metaobject_gid:
            return metaobject_gid

        if metaobject_gid in metaobject_label_cache:
            return metaobject_label_cache[metaobject_gid]

        if (
            granted_access_scopes
            and "read_metaobjects" not in granted_access_scopes
        ):
            if not warned_missing_metaobject_scope:
                print(
                    "⚠️ Current Shopify token does not include `read_metaobjects`, "
                    "so metaobject references will be exported as raw GIDs."
                )
                warned_missing_metaobject_scope = True
            metaobject_label_cache[metaobject_gid] = metaobject_gid
            return metaobject_gid

        query = """
        query MetaobjectNode($id: ID!) {
          node(id: $id) {
            ... on Metaobject {
              id
              displayName
              handle
              fields {
                key
                value
              }
            }
          }
        }
        """

        try:
            response = requests.post(
                GRAPHQL_URL,
                headers=headers,
                json={"query": query, "variables": {"id": metaobject_gid}},
                timeout=30,
            )
            response.raise_for_status()
            payload = response.json()
        except requests.RequestException as exc:
            print(f"⚠️ Failed to resolve metaobject {metaobject_gid}: {exc}")
            metaobject_label_cache[metaobject_gid] = metaobject_gid
            return metaobject_gid
        except ValueError as exc:
            print(f"⚠️ Failed to parse metaobject lookup response for {metaobject_gid}: {exc}")
            metaobject_label_cache[metaobject_gid] = metaobject_gid
            return metaobject_gid

        errors = payload.get("errors") or []
        if errors:
            print(f"⚠️ GraphQL metaobject lookup errors for {metaobject_gid}: {errors}")
            metaobject_label_cache[metaobject_gid] = metaobject_gid
            return metaobject_gid

        node = ((payload.get("data") or {}).get("node") or {})
        label_candidates = [
            node.get("displayName"),
            next(
                (
                    field.get("value")
                    for field in (node.get("fields") or [])
                    if (field.get("key") or "").strip().lower() in {"title", "name", "label"}
                    and field.get("value")
                ),
                None,
            ),
            node.get("handle"),
        ]

        resolved_label = next((candidate for candidate in label_candidates if candidate), metaobject_gid)
        metaobject_label_cache[metaobject_gid] = resolved_label
        return resolved_label

    def resolve_downloaded_metafield_value(metafield):
        field_type = (metafield.get("type") or "").strip().lower()
        raw_value = metafield.get("value")

        if not raw_value:
            return raw_value

        if field_type == "metaobject_reference":
            return get_metaobject_display_label(str(raw_value).strip())

        if field_type == "list.metaobject_reference":
            reference_values = parse_metaobject_reference_values(raw_value)
            if not reference_values:
                return raw_value
            resolved_values = [
                get_metaobject_display_label(reference_value)
                for reference_value in reference_values
            ]
            resolved_values = [value for value in resolved_values if value]
            if not resolved_values:
                return raw_value
            return ", ".join(resolved_values)

        return raw_value

    # The rest of your downloader logic goes here...

    # Fetch all locations
    def get_locations():
        url = f"{BASE_URL}/locations.json"
        response = requests.get(url, headers=headers)
        
        if response.status_code == 200:
            return response.json().get('locations', [])
        return []

    # Fetch inventory levels for a specific variant at a location
    def get_inventory_level(inventory_item_id, location_id):
        url = f"{BASE_URL}/inventory_levels.json?inventory_item_ids={inventory_item_id}&location_ids={location_id}"
        response = requests.get(url, headers=headers)
        
        if response.status_code == 200:
            inventory_levels = response.json().get('inventory_levels', [])
            if inventory_levels:
                return inventory_levels[0].get('available', 0)
        return None

    # Fetch products from Shopify using pagination
    def get_all_products():
        products = []
        url = f"{BASE_URL}/products.json?limit=250&fields=id,title,body_html,handle,tags,vendor,product_type,variants,images,created_at,updated_at,status,published_at,published_scope,template_suffix,options"

        print("Starting to fetch products...")
        page = 1  # Page counter

        while url:
            print(f"Fetching page {page} of products...")
            response = requests.get(url, headers=headers)
            if response.status_code != 200:
                print(f"Error fetching products: {response.status_code}")
                break

            batch = response.json().get('products', [])
            if not batch:  # Check if the current batch is empty to avoid unnecessary processing
                print("No more products found.")
                break

            products.extend(batch)

            # Properly handle pagination if more products are available
            link_header = response.headers.get('Link')
            next_url = None
            if link_header:
                links = link_header.split(',')
                for link in links:
                    if 'rel="next"' in link:
                        next_url = link.split(';')[0].strip('<> ').replace('&amp;', '&')
                        break

            if next_url:
                url = next_url
                page += 1
            else:
                url = None

        print(f"Finished fetching products. Total products fetched: {len(products)}")
        return products




    # Fetch metafields for a specific product
    def get_metafields(owner_id, owner_resource="product"):
        print(f"Fetching metafields for {owner_resource} ID {owner_id}...")
        time.sleep(1)  # Wait for 1 second after every request


        url = f"{BASE_URL}/metafields.json?metafield[owner_id]={owner_id}&metafield[owner_resource]={owner_resource}"
        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            return response.json().get('metafields', [])
        else:
            print(f"Error fetching metafields for {owner_resource} ID {owner_id}: {response.status_code}")
        return []

    def get_image_url_from_gid(gid):
        query = {
            "query": f"""
            {{
                media(id: "{gid}") {{
                    ... on MediaImage {{
                        image {{
                            originalSrc
                        }}
                    }}
                }}
            }}
            """
        }
        response = requests.post(GRAPHQL_URL, json=query, headers=headers)
        if response.status_code == 200:
            data = response.json()
            return data['data']['media']['image']['originalSrc']
        else:
            print(f"Failed to fetch image URL for gid {gid}: {response.status_code}")
            return None
  
    def fetch_all_metafields(products):
        print("Fetching metafields for all products concurrently...")
        product_id_to_metafields = {}
        with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
            future_to_product = {executor.submit(get_metafields, product['id']): product for product in products}
            for future in concurrent.futures.as_completed(future_to_product):
                product = future_to_product[future]
                try:
                    metafields = future.result()
                    product_id_to_metafields[product['id']] = metafields
                except Exception as exc:
                    print(f"Product ID {product['id']} generated an exception: {exc}")
                    product_id_to_metafields[product['id']] = []
        print("Finished fetching metafields.")
        return product_id_to_metafields

    def fetch_all_variant_metafields(variant_ids):
        if not variant_ids:
            return {}

        print("Fetching metafields for all variants concurrently...")
        variant_id_to_metafields = {}
        with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
            future_to_variant = {
                executor.submit(get_metafields, variant_id, "variant"): variant_id for variant_id in variant_ids
            }
            for future in concurrent.futures.as_completed(future_to_variant):
                variant_id = future_to_variant[future]
                try:
                    metafields = future.result()
                    variant_id_to_metafields[variant_id] = metafields
                except Exception as exc:
                    print(f"Variant ID {variant_id} generated an exception: {exc}")
                    variant_id_to_metafields[variant_id] = []
        print("Finished fetching variant metafields.")
        return variant_id_to_metafields

    def get_inventory_levels(inventory_item_ids):
        print("Fetching inventory levels for all variants...")
        inventory_levels = []

        # Shopify allows up to 250 inventory_item_ids per request
        batch_size = 250
        for i in range(0, len(inventory_item_ids), batch_size):
            batch_inventory_item_ids = inventory_item_ids[i:i+batch_size]
            ids_str = ','.join(map(str, batch_inventory_item_ids))
            url = f"{BASE_URL}/inventory_levels.json?inventory_item_ids={ids_str}"
            response = requests.get(url, headers=headers)
            if response.status_code == 200:
                inventory_levels.extend(response.json().get('inventory_levels', []))
            else:
                print(f"Error fetching inventory levels: {response.status_code}, {response.text}")
        print("Finished fetching inventory levels.")
        return inventory_levels

    def build_inventory_level_mapping(inventory_levels):
        inventory_mapping = {}
        for level in inventory_levels:
            key = (level['inventory_item_id'], level['location_id'])
            inventory_mapping[key] = level['available']
        return inventory_mapping

    # Main function to save product, variant, and inventory data to Excel
    def save_to_excel(products, locations):
        print("Processing data and preparing to save to Excel...")
        data = []
        all_metafield_keys = set()
        max_image_columns = 0  # Track the maximum number of images

         # Collect all inventory_item_ids
        inventory_item_ids = set()
        variant_ids = []
        for product in products:
            images = product.get('images', [])
            max_image_columns = max(max_image_columns, len(images))
            for variant in product.get('variants', []):
                inventory_item_id = variant.get('inventory_item_id')
                if inventory_item_id:
                    inventory_item_ids.add(inventory_item_id)
                variant_ids.append(variant['id'])



        # Fetch inventory levels for all inventory items
        inventory_levels = get_inventory_levels(list(inventory_item_ids))
        inventory_mapping = build_inventory_level_mapping(inventory_levels)

        # Fetch metafields for all products concurrently
        product_id_to_metafields = fetch_all_metafields(products)
        variant_id_to_metafields = fetch_all_variant_metafields(list(set(variant_ids)))

        # Second pass to add data, including images and inventory levels
        product_count = len(products)
        for idx, product in enumerate(products, 1):
            print(f"Processing product {idx}/{product_count}: {product['title']}")

            # Ensure that tags are correctly handled and joined as strings
            tags = product.get('tags', [])
            if isinstance(tags, str):
                tags = tags.replace(', ', ',').split(',')
            if isinstance(tags, list):
                tags = ', '.join(tags)
            else:
                tags = str(tags)

            # Build image mapping from image ID to image details (URL and alt)
            images = product.get('images', [])
            image_id_to_image = {}
            for image in images:
                image_id = image.get('id')
                if image_id:
                    image_id_to_image[image_id] = {
                        'src': image.get('src'),
                        'alt': image.get('alt', '')
                    }

            # First product row (with all product-level details)
            first_variant = product.get('variants', [])[0]  # The first variant to display in the product row
            variant_id = first_variant.get('id', '')

            # Get variant image details if available
            variant_image_url = ''
            variant_image_alt = ''
            image_id = first_variant.get('image_id')
            if image_id and image_id in image_id_to_image:
                variant_image_url = image_id_to_image[image_id]['src']
                variant_image_alt = image_id_to_image[image_id]['alt']



            product_data = {
                "Title": product['title'],
                "Handle": product['handle'],
                "ID": product['id'],
                "Body HTML": product.get('body_html', ''),
                "Vendor": product['vendor'],
                "Type": product.get('product_type', ''),
                "Tags": tags,
                "Created At": product['created_at'],
                "Updated At": product['updated_at'],
                "Status": product['status'],
                "Published": product['published_at'],
                "Published Scope": product.get('published_scope', ''),
                "Template Suffix": product.get('template_suffix', ''),
                "Variant ID": first_variant.get('id', ''),
                "Option1 Name": product.get('options', [{}])[0].get('name', '') if len(product.get('options', [])) > 0 else "",
                "Option1 Value": first_variant.get('option1', '') if len(product.get('options', [])) > 0 else "",
                "Option2 Name": product.get('options', [{}])[1].get('name', '') if len(product.get('options', [])) > 1 else "",
                "Option2 Value": first_variant.get('option2', '') if len(product.get('options', [])) > 1 else "",
                "Option3 Name": product.get('options', [{}])[2].get('name', '') if len(product.get('options', [])) > 2 else "",
                "Option3 Value": first_variant.get('option3', '') if len(product.get('options', [])) > 2 else "",
                "Variant SKU": first_variant.get('sku', ''),
                "Variant Price": first_variant.get('price', ''),  
                "Variant Compare At Price": first_variant.get('compare_at_price', ''),
                "Variant Inventory Qty": first_variant.get('inventory_quantity', 0),
                "Variant Weight": first_variant.get('weight', ''),
                "Variant Weight Unit": first_variant.get('weight_unit', ''),
                "Variant Barcode": first_variant.get('barcode', ''),
                "Continue Selling When Sold Out": first_variant.get('inventory_policy', ''),
                "Variant Image": variant_image_url,
                "Variant Image Alt": variant_image_alt,
                "Status": product['status'],  # Include status directly in the product data

            }

            # Add image URLs and alt texts dynamically after the variant data
            for i in range(max_image_columns):
                if i < len(images):
                    image_url = images[i]['src']
                    image_alt = images[i].get('alt', '')
                else:
                    image_url = None
                    image_alt = None
                product_data[f"Image {i + 1}"] = image_url
                product_data[f"Image {i + 1} Alt"] = image_alt

            
             # Add product-level metafields after the images
            metafields = product_id_to_metafields.get(product['id'], [])
            for metafield in metafields:
                key = metafield['key']
                value = resolve_downloaded_metafield_value(metafield)
                namespace = metafield['namespace']
                field_type = metafield.get('type', 'unknown')
                column_name = f"Metafield: {namespace}.{key} [{field_type}]"
                all_metafield_keys.add(column_name)
                product_data[column_name] = value

            # Add variant-level metafields for the first variant
            first_variant_metafields = variant_id_to_metafields.get(first_variant.get('id'), [])
            for metafield in first_variant_metafields:
                key = metafield['key']
                value = resolve_downloaded_metafield_value(metafield)
                namespace = metafield['namespace']
                field_type = metafield.get('type', 'unknown')
                column_name = f"Variant Metafield: {namespace}.{key} [{field_type}]"
                all_metafield_keys.add(column_name)
                product_data[column_name] = value

            # Get inventory levels for each location for the first variant
            for location in locations:
                location_id = location['id']
                location_name = location['name']
                key = (first_variant.get('inventory_item_id'), location_id)
                inventory_level = inventory_mapping.get(key, 0)
                product_data[f"Inventory Available: {location_name}"] = inventory_level

            # Append the first row for the product
            data.append(product_data)

            # Additional rows for the rest of the variants (without product-level details)
            variant_count = len(product.get('variants', []))
            for v_idx, variant in enumerate(product.get('variants', [])[1:], 2):
                print(f"  Processing variant {v_idx}/{variant_count}: {variant.get('sku', '')}")
                variant_id = variant.get('id', '')
                variant_image_url = ''
                variant_image_alt = ''
                image_id = variant.get('image_id')
                if image_id and image_id in image_id_to_image:
                    variant_image_url = image_id_to_image[image_id]['src']
                    variant_image_alt = image_id_to_image[image_id]['alt']

                variant_data = {
                    "ID": "",  # Leave product ID blank for variants
                    "Variant ID": variant.get('id', ''),
                    "Option1 Name": product.get('options', [{}])[0].get('name', '') if len(product.get('options', [])) > 0 else "",
                    "Option1 Value": variant.get('option1', '') if len(product.get('options', [])) > 0 else "",
                    "Option2 Name": product.get('options', [{}])[1].get('name', '') if len(product.get('options', [])) > 1 else "",
                    "Option2 Value": variant.get('option2', '') if len(product.get('options', [])) > 1 else "",
                    "Option3 Name": product.get('options', [{}])[2].get('name', '') if len(product.get('options', [])) > 2 else "",
                    "Option3 Value": variant.get('option3', '') if len(product.get('options', [])) > 2 else "",
                    "Variant SKU": variant.get('sku', ''),
                    "Variant Price": variant.get('price', ''),
                    "Variant Compare At Price": variant.get('compare_at_price', ''),
                    "Variant Inventory Qty": variant.get('inventory_quantity', 0),
                    "Variant Weight": variant.get('weight', ''),
                    "Variant Weight Unit": variant.get('weight_unit', ''),
                    "Variant Barcode": variant.get('barcode', ''),
                    "Continue Selling When Sold Out": variant.get('inventory_policy', ''),
                    "Variant Image": variant_image_url,
                    "Variant Image Alt": variant_image_alt
                }

                # Get inventory levels for each location
                for location in locations:
                    location_id = location['id']
                    location_name = location['name']
                    key = (variant.get('inventory_item_id'), location_id)
                    inventory_level = inventory_mapping.get(key, 0)
                    variant_data[f"Inventory Available: {location_name}"] = inventory_level

                variant_metafields = variant_id_to_metafields.get(variant.get('id'), [])
                for metafield in variant_metafields:
                    key = metafield['key']
                    value = resolve_downloaded_metafield_value(metafield)
                    namespace = metafield['namespace']
                    field_type = metafield.get('type', 'unknown')
                    column_name = f"Variant Metafield: {namespace}.{key} [{field_type}]"
                    all_metafield_keys.add(column_name)
                    variant_data[column_name] = value

                # Append the variant row
                data.append(variant_data)

            # Append a blank row for separation (optional)
            data.append({})  # Adding an empty row for grouping in Excel

        # Create DataFrame and ensure all metafield columns are present
        df = pd.DataFrame(data)
        for key in all_metafield_keys:
            if key not in df.columns:
                df[key] = None

        # Get the current date and time to append to the filename
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = f"shopify_products_bulk_{current_time}.xlsx"

        # Save to Excel
        df.to_excel(file_path, index=False)

        # Open the saved Excel file using openpyxl to apply freezing panes and bold styling
        wb = load_workbook(file_path)
        ws = wb.active

        # Apply bold to product rows (rows where "ID" is present)
        bold_font = Font(bold=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[2].value:  # Assuming "ID" is in the 3rd column (0-based index)
                for cell in row:
                    cell.font = bold_font

        # Freeze the first row and first column
        ws.freeze_panes = ws['B2']  # Freeze the first row and column

        # Save the changes
        wb.save(file_path)

        print("Data has been saved successfully.")

    # Example usage
    products = get_all_products()
    locations = get_locations()
    if products and locations:
        save_to_excel(products, locations)

    print(f"Data has been saved.")

def run_uploader_logic():
    # Get the directory where the executable or script is located
    if getattr(sys, 'frozen', False):  # If running as an EXE
        script_dir = os.path.dirname(sys.executable)
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))

    # Build the full path to 'credentials.txt'
    credentials_path = os.path.join(script_dir, 'credentials.txt')

    shopify_context = load_shopify_context(credentials_path)
    SHOP_NAME = shopify_context["shop_name"]
    ACCESS_TOKEN = shopify_context["access_token"]
    log_shopify_access_scope_diagnostics(
        shopify_context,
        required_scopes=["read_metaobjects", "read_metaobject_definitions"],
    )
    granted_access_scopes = set(shopify_context.get("granted_access_scopes") or [])

    # Shopify API URLs
    BASE_URL = shopify_context["base_url"]
    GRAPHQL_URL = shopify_context["graphql_url"]

    # Headers for API authentication
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": ACCESS_TOKEN
    }
    graphql_headers = headers.copy()

    global append_media_to_variant, get_product_id_by_handle

    def _get_media_status(media_gid):
        """Return the processing status for a Media GID."""
        query = """
        query MediaStatus($id: ID!) {
          node(id: $id) {
            __typename
            ... on MediaImage {
              id
              status
            }
            ... on Video {
              id
              status
            }
            ... on Model3d {
              id
              status
            }
          }
        }
        """
        data = _graphql_post(query, {"id": media_gid}, purpose="Media status")
        node = (((data or {}).get("data") or {}).get("node") or {})
        return node.get("status")

    def _wait_until_media_ready(media_gid, timeout_s=240, interval_s=3):
        """Poll the media until Shopify reports it as READY."""
        import time as _t

        deadline = _t.time() + timeout_s
        while _t.time() < deadline:
            status = _get_media_status(media_gid)
            if status == "READY":
                return True
            if status == "FAILED":
                print(f"❌ Media {media_gid} failed processing; cannot attach to variant.")
                return False
            _t.sleep(interval_s)
        print(f"⚠️ Timed out waiting for media {media_gid} to become READY.")
        return False


    def _graphql_post(query, variables=None, purpose="(unspecified)"):
        payload = {"query": query, "variables": variables or {}}
        try:
            resp = requests.post(GRAPHQL_URL, headers=graphql_headers, json=payload, timeout=30)
        except Exception as e:
            print(f"❌ GraphQL {purpose} request failed to send: {e}")
            return None

        # Try to parse JSON, log raw text when impossible
        try:
            data = resp.json()
        except Exception:
            print(f"❌ GraphQL {purpose} non-JSON response. HTTP {resp.status_code}. Body:\n{resp.text[:800]}")
            return None

        # Shopify GraphQL may attach top-level "errors"
        if "errors" in data:
            print(f"❌ GraphQL {purpose} errors: {data['errors']}")
        return data

    metaobject_definition_cache = {}
    metaobject_lookup_cache = {}
    metaobject_known_reference_cache = {}
    owner_summary_cache = {}

    def normalize_metaobject_lookup_value(value):
        if value is None:
            return ""
        return re.sub(r"\s+", " ", str(value)).strip().casefold()

    def parse_metaobject_reference_values(value):
        if value is None:
            return []

        if isinstance(value, (list, tuple, set)):
            return [str(item).strip() for item in value if str(item).strip()]

        try:
            if pd.isna(value):
                return []
        except TypeError:
            pass

        raw_value = str(value).strip()
        if not raw_value:
            return []

        if raw_value.startswith("["):
            try:
                parsed = json.loads(raw_value)
            except json.JSONDecodeError:
                parsed = None
            if isinstance(parsed, list):
                return [str(item).strip() for item in parsed if str(item).strip()]

        return [item.strip() for item in re.split(r"[\n,;]+", raw_value) if item.strip()]

    def get_metafield_metaobject_definition(owner_type, namespace, key):
        cache_key = (owner_type, namespace, key)
        if cache_key in metaobject_definition_cache:
            return metaobject_definition_cache[cache_key]

        query = """
        query MetafieldDefinitionLookup(
          $ownerType: MetafieldOwnerType!,
          $namespace: String!,
          $key: String!
        ) {
          metafieldDefinitions(
            first: 1,
            ownerType: $ownerType,
            namespace: $namespace,
            key: $key
          ) {
            nodes {
              id
              name
              validations {
                name
                value
              }
            }
          }
        }
        """
        data = _graphql_post(
            query,
            {
                "ownerType": owner_type,
                "namespace": namespace,
                "key": key,
            },
            purpose=f"Metafield definition lookup for {owner_type} {namespace}.{key}",
        )
        definition_nodes = (((data or {}).get("data") or {}).get("metafieldDefinitions") or {}).get("nodes") or []
        if not definition_nodes:
            print(
                f"⚠️ Could not find a metafield definition for {owner_type} {namespace}.{key}; "
                "metaobject titles cannot be resolved."
            )
            metaobject_definition_cache[cache_key] = None
            return None

        definition_node = definition_nodes[0]
        definition_gid = None
        metaobject_type = None

        for validation in definition_node.get("validations") or []:
            validation_name = (validation.get("name") or "").strip().lower()
            if "metaobject_definition" not in validation_name:
                continue

            for candidate in parse_metaobject_reference_values(validation.get("value")):
                if candidate.startswith("gid://shopify/MetaobjectDefinition/"):
                    definition_gid = candidate
                    break
                if candidate and not metaobject_type:
                    metaobject_type = candidate

            if definition_gid:
                break

        definition_name = definition_node.get("name")
        if definition_gid:
            definition_query = """
            query MetaobjectDefinitionLookup($id: ID!) {
              node(id: $id) {
                ... on MetaobjectDefinition {
                  id
                  name
                  type
                }
              }
            }
            """
            definition_data = _graphql_post(
                definition_query,
                {"id": definition_gid},
                purpose=f"Metaobject definition details for {namespace}.{key}",
            )
            definition_details = (((definition_data or {}).get("data") or {}).get("node") or {})
            definition_name = definition_details.get("name") or definition_name
            metaobject_type = definition_details.get("type") or metaobject_type

        if not metaobject_type:
            if granted_access_scopes and "read_metaobject_definitions" not in granted_access_scopes:
                print(
                    "⚠️ The current Shopify token does not include "
                    "`read_metaobject_definitions`, so metaobject titles cannot be "
                    "resolved by definition lookup yet."
                )
            print(
                f"⚠️ Found metafield definition for {namespace}.{key}, but could not determine "
                "which metaobject type it references."
            )
            metaobject_definition_cache[cache_key] = None
            return None

        definition_info = {
            "definition_id": definition_gid,
            "definition_name": definition_name,
            "metaobject_type": metaobject_type,
        }
        metaobject_definition_cache[cache_key] = definition_info
        return definition_info

    def add_metaobject_lookup_candidate(lookup, raw_value, entry):
        normalized_value = normalize_metaobject_lookup_value(raw_value)
        if not normalized_value:
            return

        matches = lookup.setdefault(normalized_value, [])
        if all(existing["id"] != entry["id"] for existing in matches):
            matches.append(entry)

    def get_metaobject_lookup(metaobject_type):
        if metaobject_type in metaobject_lookup_cache:
            return metaobject_lookup_cache[metaobject_type]

        query = """
        query MetaobjectsByType($type: String!, $after: String) {
          metaobjects(type: $type, first: 250, after: $after) {
            nodes {
              id
              displayName
              handle
              fields {
                key
                value
              }
            }
            pageInfo {
              hasNextPage
              endCursor
            }
          }
        }
        """

        lookup = {}
        has_next_page = True
        cursor = None

        while has_next_page:
            data = _graphql_post(
                query,
                {"type": metaobject_type, "after": cursor},
                purpose=f"Metaobjects lookup for type {metaobject_type}",
            )
            connection = (((data or {}).get("data") or {}).get("metaobjects") or {})
            nodes = connection.get("nodes") or []

            for node in nodes:
                metaobject_id = node.get("id")
                if not metaobject_id:
                    continue

                entry = {
                    "id": metaobject_id,
                    "display_name": node.get("displayName") or "",
                    "handle": node.get("handle") or "",
                }
                add_metaobject_lookup_candidate(lookup, entry["display_name"], entry)
                add_metaobject_lookup_candidate(lookup, entry["handle"], entry)

                for field in node.get("fields") or []:
                    field_key = (field.get("key") or "").strip().lower()
                    field_value = field.get("value")
                    if field_key in {"title", "name", "label"} and field_value:
                        add_metaobject_lookup_candidate(lookup, field_value, entry)

            page_info = connection.get("pageInfo") or {}
            has_next_page = bool(page_info.get("hasNextPage"))
            cursor = page_info.get("endCursor")
            if has_next_page and not cursor:
                has_next_page = False

        metaobject_lookup_cache[metaobject_type] = lookup
        return lookup

    def get_owner_summaries_for_metaobject_debug(owner_type):
        if owner_type in owner_summary_cache:
            return owner_summary_cache[owner_type]

        owner_summaries = []
        for product in get_all_products_forsku() or []:
            product_id = product.get("id")
            product_title = product.get("title") or ""
            product_handle = product.get("handle") or ""

            if owner_type == "PRODUCT":
                if product_id:
                    owner_summaries.append(
                        {
                            "owner_id": str(product_id),
                            "owner_label": product_handle or product_title or str(product_id),
                            "product_title": product_title,
                            "product_handle": product_handle,
                        }
                    )
                continue

            if owner_type == "PRODUCTVARIANT":
                for variant in product.get("variants", []) or []:
                    variant_id = variant.get("id")
                    if not variant_id:
                        continue
                    variant_sku = variant.get("sku") or ""
                    owner_summaries.append(
                        {
                            "owner_id": str(variant_id),
                            "owner_label": variant_sku or f"{product_handle}:{variant_id}",
                            "product_title": product_title,
                            "product_handle": product_handle,
                            "variant_sku": variant_sku,
                        }
                    )

        owner_summary_cache[owner_type] = owner_summaries
        return owner_summaries

    def fetch_metaobject_node_by_gid(metaobject_gid):
        query = """
        query MetaobjectNode($id: ID!) {
          node(id: $id) {
            ... on Metaobject {
              id
              displayName
              handle
              type
              fields {
                key
                value
              }
            }
          }
        }
        """
        data = _graphql_post(
            query,
            {"id": metaobject_gid},
            purpose=f"Metaobject lookup for {metaobject_gid}",
        )
        return (((data or {}).get("data") or {}).get("node") or {})

    def build_known_metaobject_reference_cache(owner_type, namespace, key):
        cache_key = (owner_type, namespace, key)
        if cache_key in metaobject_known_reference_cache:
            return metaobject_known_reference_cache[cache_key]

        endpoint = "products" if owner_type == "PRODUCT" else "variants"
        owner_summaries = get_owner_summaries_for_metaobject_debug(owner_type)
        metaobject_references = {}

        print(
            f"🔎 Building known metaobject reference list for {owner_type} {namespace}.{key} "
            f"from {len(owner_summaries)} owners..."
        )

        for owner_summary in owner_summaries:
            owner_id = owner_summary["owner_id"]
            try:
                response = requests.get(
                    f"{BASE_URL}/{endpoint}/{owner_id}/metafields.json",
                    headers=headers,
                    params={"limit": 250},
                    timeout=30,
                )
            except requests.RequestException as exc:
                print(f"⚠️ Failed to fetch metafields for {endpoint[:-1]} {owner_id}: {exc}")
                continue

            if response.status_code != 200:
                print(
                    f"⚠️ Failed to fetch metafields for {endpoint[:-1]} {owner_id}: "
                    f"{response.status_code}, {response.text[:300]}"
                )
                continue

            for metafield in response.json().get("metafields", []) or []:
                if metafield.get("namespace") != namespace or metafield.get("key") != key:
                    continue
                for candidate_gid in parse_metaobject_reference_values(metafield.get("value")):
                    if not candidate_gid.startswith("gid://shopify/Metaobject/"):
                        continue
                    references = metaobject_references.setdefault(candidate_gid, [])
                    owner_label = owner_summary.get("owner_label") or owner_id
                    if owner_label not in references:
                        references.append(owner_label)

        lookup = {}
        debug_rows = []

        for metaobject_gid, references in metaobject_references.items():
            node = fetch_metaobject_node_by_gid(metaobject_gid)
            display_name = node.get("displayName") or ""
            handle_value = node.get("handle") or ""
            metaobject_type = node.get("type") or ""

            entry = {
                "id": metaobject_gid,
                "display_name": display_name,
                "handle": handle_value,
                "type": metaobject_type,
                "references": references,
            }

            labels = []
            for label_candidate in [display_name, handle_value]:
                if label_candidate and label_candidate not in labels:
                    labels.append(label_candidate)

            for field in node.get("fields") or []:
                field_key = (field.get("key") or "").strip().lower()
                field_value = (field.get("value") or "").strip()
                if field_key in {"title", "name", "label"} and field_value and field_value not in labels:
                    labels.append(field_value)

            if not labels:
                labels.append(metaobject_gid)

            for label in labels:
                add_metaobject_lookup_candidate(lookup, label, entry)

            debug_rows.append(
                {
                    "Metaobject ID": metaobject_gid,
                    "Display Name": display_name,
                    "Handle": handle_value,
                    "Type": metaobject_type,
                    "Known Labels": " | ".join(labels),
                    "Referenced By": " ; ".join(references),
                }
            )

        export_path = None
        if debug_rows:
            safe_name = re.sub(r"[^a-zA-Z0-9._-]+", "_", f"{namespace}_{key}")
            export_path = os.path.join(
                script_dir,
                f"metaobject_debug_{safe_name}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx",
            )
            try:
                pd.DataFrame(debug_rows).to_excel(export_path, index=False)
                print(
                    f"📝 Wrote known metaobject reference list for {namespace}.{key} to "
                    f"{export_path}"
                )
            except Exception as exc:
                print(f"⚠️ Failed to write metaobject debug file for {namespace}.{key}: {exc}")
                export_path = None
        else:
            print(f"⚠️ No existing metaobject references found for {namespace}.{key} in the store.")

        cache_value = {
            "lookup": lookup,
            "debug_rows": debug_rows,
            "export_path": export_path,
        }
        metaobject_known_reference_cache[cache_key] = cache_value
        return cache_value

    def resolve_metaobject_from_known_references(owner_type, namespace, key, raw_value):
        known_reference_cache = build_known_metaobject_reference_cache(owner_type, namespace, key)
        lookup = known_reference_cache.get("lookup", {})
        matches = lookup.get(normalize_metaobject_lookup_value(raw_value), [])

        if len(matches) == 1:
            resolved_gid = matches[0]["id"]
            print(
                f"🔗 Resolved metaobject '{raw_value}' from known {namespace}.{key} "
                f"references: {resolved_gid}"
            )
            return resolved_gid

        debug_path = known_reference_cache.get("export_path")
        if len(matches) > 1:
            options = ", ".join(
                match.get("display_name") or match.get("handle") or match.get("id")
                for match in matches
            )
            print(
                f"❌ Metaobject value '{raw_value}' is ambiguous in known {namespace}.{key} "
                f"references. Matches: {options}"
            )
            if debug_path:
                print(f"📝 Review the debug list here: {debug_path}")
            return None

        if debug_path:
            print(
                f"⚠️ '{raw_value}' was not found in the known {namespace}.{key} reference list. "
                f"Review the debug list here: {debug_path}"
            )
        return None

    def resolve_metaobject_reference_gid(owner_type, namespace, key, raw_value):
        if isinstance(raw_value, str):
            raw_value = raw_value.strip()

        if not raw_value:
            return None

        if isinstance(raw_value, str) and raw_value.startswith("gid://shopify/Metaobject/"):
            return raw_value

        definition_info = get_metafield_metaobject_definition(owner_type, namespace, key)
        if not definition_info:
            return resolve_metaobject_from_known_references(owner_type, namespace, key, raw_value)

        lookup = get_metaobject_lookup(definition_info["metaobject_type"])
        matches = lookup.get(normalize_metaobject_lookup_value(raw_value), [])

        if len(matches) == 1:
            resolved_gid = matches[0]["id"]
            print(
                f"🔗 Resolved metaobject '{raw_value}' to {resolved_gid} for "
                f"{namespace}.{key}"
            )
            return resolved_gid

        if len(matches) > 1:
            options = ", ".join(
                match["display_name"] or match["handle"] or match["id"] for match in matches
            )
            print(
                f"❌ Metaobject value '{raw_value}' is ambiguous for {namespace}.{key}. "
                f"Matches: {options}"
            )
            return resolve_metaobject_from_known_references(owner_type, namespace, key, raw_value)

        print(
            f"❌ Could not resolve metaobject value '{raw_value}' for {namespace}.{key}. "
            f"Expected a metaobject of type '{definition_info['metaobject_type']}'."
        )
        return resolve_metaobject_from_known_references(owner_type, namespace, key, raw_value)

    def resolve_metaobject_reference_value(owner_type, namespace, key, value):
        values = parse_metaobject_reference_values(value)
        if not values:
            return None

        if len(values) > 1:
            print(
                f"❌ Metafield {namespace}.{key} accepts a single metaobject reference, "
                f"but received {len(values)} values: {values}"
            )
            return None

        return resolve_metaobject_reference_gid(owner_type, namespace, key, values[0])

    def resolve_metaobject_reference_list(owner_type, namespace, key, value):
        resolved_gids = []
        seen = set()

        for raw_item in parse_metaobject_reference_values(value):
            resolved_gid = resolve_metaobject_reference_gid(owner_type, namespace, key, raw_item)
            if not resolved_gid:
                return None
            if resolved_gid not in seen:
                seen.add(resolved_gid)
                resolved_gids.append(resolved_gid)

        return resolved_gids

    def to_gid(resource, numeric_id):
        return f"gid://shopify/{resource}/{numeric_id}"

    def get_product_id_by_handle(handle):
        url = f"{BASE_URL}/products.json?handle={handle}"
        r = requests.get(url, headers=headers)
        if r.status_code == 200 and r.json().get("products"):
            return r.json()["products"][0]["id"]
        print(f"⚠️ Could not find product by handle '{handle}'. HTTP {r.status_code} Body: {r.text[:400]}")
        return None

    def _get_product_image_info(product_id, product_image_gid):
        """
        From a Product + ProductImage GID, fetch URL and alt text for that image.
        Returns dict: {"url": str, "alt": str} or None.
        """
        query = """
        query ProductImages($id: ID!) {
          product(id: $id) {
            id
            images(first: 250) {
              nodes {
                id
                url
                altText
              }
            }
          }
        }
        """
        product_gid = to_gid("Product", product_id)
        data = _graphql_post(query, {"id": product_gid}, purpose="ProductImages lookup")
        if not data:
            return None

        nodes = ((((data or {}).get("data") or {})
                    .get("product") or {})
                    .get("images", {})
                    .get("nodes", []))
        for node in nodes:
            if node.get("id") == product_image_gid:
                return {"url": node.get("url"), "alt": node.get("altText")}

        print(f"⚠️ Did not find ProductImage {product_image_gid} in product images list.")
        return None

    def _create_mediaimage_from_productimage(product_id, image_url, alt_text=None):
        """
        Calls productCreateMedia to create a MediaImage from a raw image URL.
        Returns the new MediaImage gid or None.
        """
        mutation = """
        mutation CreateMedia($productId: ID!, $media: [CreateMediaInput!]!) {
          productCreateMedia(productId: $productId, media: $media) {
            media {
              __typename
              ... on MediaImage {
                id
                image { id }
              }
            }
            mediaUserErrors { field message code }
          }
        }
        """
        variables = {
            "productId": to_gid("Product", product_id),
            "media": [{
                "mediaContentType": "IMAGE",  # <-- REQUIRED
                "originalSource": image_url,
                "alt": alt_text or ""
            }]
        }
        data = _graphql_post(mutation, variables, purpose="productCreateMedia")
        if not data:
            print("❌ No data returned from productCreateMedia.")
            return None

        errs = ((((data or {}).get("data") or {})
                    .get("productCreateMedia") or {})
                    .get("mediaUserErrors", []))
        if errs:
            print(f"❌ productCreateMedia errors: {errs}")
            return None

        media_nodes = ((((data or {}).get("data") or {})
                           .get("productCreateMedia") or {})
                           .get("media", []))
        for node in media_nodes:
            if node.get("__typename") == "MediaImage":
                media_gid = node.get("id")
                if not media_gid:
                    continue
                if not _wait_until_media_ready(media_gid, timeout_s=240, interval_s=3):
                    return None
                return media_gid

        print("⚠️ productCreateMedia returned no MediaImage nodes.")
        return None

    def _resolve_mediaimage_id_from_productimage_gid(product_gid, product_image_gid):
        """
        Map legacy ProductImage GID -> MediaImage GID by querying product.media and matching image.id.
        Returns a MediaImage gid or None (does NOT create).
        """
        query = """
        query ProductMedia($id: ID!) {
          product(id: $id) {
            id
            media(first: 250) {
              nodes {
                __typename
                ... on MediaImage {
                  id
                  image { id }
                }
              }
            }
          }
        }
        """
        data = _graphql_post(query, {"id": product_gid}, purpose="ProductMedia lookup")
        if not data:
            return None

        nodes = ((((data or {}).get("data") or {})
                    .get("product") or {})
                    .get("media", {})
                    .get("nodes", []))
        for node in nodes:
            if node.get("__typename") == "MediaImage":
                image = node.get("image") or {}
                if image.get("id") == product_image_gid:
                    return node.get("id")
        return None

    def _ensure_media_gid_for_append(product_id, incoming_gid):
        """
        Ensure we pass a Media gid to productVariantAppendMedia.
        If we receive a ProductImage gid, try to map. If mapping fails, create a MediaImage from that ProductImage.
        """
        if incoming_gid and incoming_gid.startswith("gid://shopify/Media"):
            return incoming_gid  # already a media gid

        if incoming_gid and incoming_gid.startswith("gid://shopify/ProductImage/"):
            product_gid = to_gid("Product", product_id)

            # 1) Try to map an existing MediaImage
            mapped = _resolve_mediaimage_id_from_productimage_gid(product_gid, incoming_gid)
            if mapped:
                return mapped

            print(f"⚠️ No MediaImage found that corresponds to ProductImage {incoming_gid}; creating one via productCreateMedia...")

            # 2) Create MediaImage from ProductImage (needs image URL)
            info = _get_product_image_info(product_id, incoming_gid)
            if not info or not info.get("url"):
                print(f"❌ Could not fetch URL for {incoming_gid}; cannot create MediaImage.")
                return None

            created_media_gid = _create_mediaimage_from_productimage(product_id, info["url"], info.get("alt"))
            if created_media_gid:
                return created_media_gid

            return None

        print(f"⚠️ Unsupported media identifier (expect Media* or ProductImage gid): {incoming_gid}")
        return None

    def append_media_to_variant(product_id, variant_id, incoming_media_gid):
        media_gid = _ensure_media_gid_for_append(product_id, incoming_media_gid)
        if not media_gid:
            print(f"❌ Cannot append to variant {variant_id}: failed to resolve a Media gid from {incoming_media_gid}")
            return

        # Ensure the media is processed before attempting to attach it
        if not _wait_until_media_ready(media_gid, timeout_s=240, interval_s=3):
            print(
                f"❌ Media {media_gid} was not ready in time; skipping append for variant {variant_id}."
            )
            return

        mutation = """
        mutation AppendVariantMedia($productId: ID!, $variantMedia: [ProductVariantAppendMediaInput!]!) {
          productVariantAppendMedia(productId: $productId, variantMedia: $variantMedia) {
            userErrors { field message code }
          }
        }
        """
        variables = {
            "productId": to_gid("Product", product_id),
            "variantMedia": [{
                "variantId": to_gid("ProductVariant", variant_id),
                "mediaIds": [media_gid]
            }]
        }

        data = _graphql_post(mutation, variables, purpose="productVariantAppendMedia")
        if not data:
            print(f"❌ No data returned from productVariantAppendMedia for variant {variant_id}")
            return

        user_errors = ((((data or {}).get("data") or {})
                           .get("productVariantAppendMedia") or {})
                           .get("userErrors", []))
        if user_errors:
            # Optional: auto-retry once if NON_READY_MEDIA somehow sneaks through
            if any(err.get("code") == "NON_READY_MEDIA" for err in user_errors):
                print("⏳ Media not ready at append time; waiting a bit and retrying once...")
                if not _wait_until_media_ready(media_gid, timeout_s=180, interval_s=3):
                    print(
                        f"❌ Media {media_gid} failed to become ready after retry wait; "
                        f"skipping append for variant {variant_id}."
                    )
                    return
                data = _graphql_post(mutation, variables, purpose="productVariantAppendMedia (retry)")
                user_errors = ((((data or {}).get("data") or {})
                                   .get("productVariantAppendMedia") or {})
                                   .get("userErrors", []))
            if user_errors:
                print(f"❌ productVariantAppendMedia errors for variant {variant_id}: {user_errors}")
            else:
                print(f"✅ Appended media {media_gid} to variant {variant_id}")
        else:
            print(f"✅ Appended media {media_gid} to variant {variant_id}")


    # Image folder path
    IMAGE_FOLDER = os.path.join(script_dir, 'img')  # Adjusted to your images folder
    FILE_FOLDER = os.path.join(script_dir, 'files')  # Optional folder for non-image assets

    asset_directories = []
    for candidate in [IMAGE_FOLDER, FILE_FOLDER, script_dir]:
        if candidate and candidate not in asset_directories:
            asset_directories.append(candidate)

    def resolve_local_asset_path(filename):
        return resolve_asset_from_directories(filename, asset_directories)

    def fetch_primary_location_id():
        """Retrieve the first active Shopify location for inventory updates."""
        url = f"{BASE_URL}/locations.json"
        try:
            response = requests.get(url, headers=headers)
        except requests.RequestException as exc:
            print(f"Failed to fetch locations due to network error: {exc}")
            return None

        if response.status_code != 200:
            print(f"Failed to fetch locations: {response.status_code}, {response.text}")
            return None

        locations = response.json().get("locations", [])
        if not locations:
            print("No Shopify locations available to update inventory levels.")
            return None

        # Prefer active locations; fall back to the first entry if none are marked active
        active_locations = [loc for loc in locations if loc.get("active", True)]
        selected_location = (active_locations or locations)[0]

        location_id = selected_location.get("id")
        location_name = selected_location.get("name", "Unknown")
        print(f"Using location '{location_name}' (ID: {location_id}) for inventory updates.")
        return location_id

    primary_location_id = fetch_primary_location_id()

    def set_inventory_level(inventory_item_id, quantity):
        """Synchronize the available inventory for an item at the primary location."""
        if inventory_item_id is None:
            print("Cannot update inventory: missing inventory_item_id.")
            return

        if primary_location_id is None:
            print("Cannot update inventory: no Shopify location available.")
            return

        if quantity is None:
            print("Inventory quantity is None; skipping inventory update.")
            return

        try:
            available_quantity = int(float(quantity))
        except (TypeError, ValueError):
            print(f"Invalid inventory quantity '{quantity}' provided; skipping inventory update.")
            return

        payload = {
            "location_id": primary_location_id,
            "inventory_item_id": inventory_item_id,
            "available": available_quantity
        }

        try:
            response = requests.post(
                f"{BASE_URL}/inventory_levels/set.json",
                headers=headers,
                json=payload
            )
        except requests.RequestException as exc:
            print(f"Network error while setting inventory level: {exc}")
            return

        if response.status_code in (200, 201):
            print(
                f"Inventory updated for item {inventory_item_id} at location {primary_location_id} to {quantity}."
            )
        else:
            print(
                f"Failed to set inventory level for item {inventory_item_id}: "
                f"{response.status_code}, {response.text}"
            )

    # Function to clean data by removing NaN values and converting them to None
    def clean_data(data):
        """
        Recursively clean data to replace NaN, inf, -inf with None, and handle invalid values for JSON serialization.
        """
        if isinstance(data, dict):
            return {k: clean_data(v) for k, v in data.items()}
        elif isinstance(data, list):
            return [clean_data(v) for v in data]
        elif isinstance(data, float):
            # Replace NaN, inf, -inf with None
            if pd.isna(data) or data in [float('inf'), float('-inf')]:
                return None
        return data

    




    # Function to update product images
    def update_product_images(product_id, images, alt_texts):
        product_id = str(int(product_id)) if not pd.isna(product_id) else None
        if not product_id:
            print(f"Invalid product ID, skipping image update.")
            return

        # Build the image data list including alt texts
        image_data_list = []
        for image_url, alt_text in zip(images, alt_texts):
            if pd.notna(image_url) and image_url:
                image_data = {"src": image_url}
                if pd.notna(alt_text) and alt_text:
                    image_data["alt"] = alt_text
                image_data_list.append(image_data)

        if not image_data_list:
            print(f"No valid images found for product {product_id}, skipping.")
            return

        url = f"{BASE_URL}/products/{product_id}.json"
        product_data = {
            "product": {
                "id": product_id,
                "images": image_data_list,
                "alt": alt_text

            }
        }

        try:
            response = requests.put(url, headers=headers, json=product_data)
            response.raise_for_status()  # Raise an exception for HTTP errors
            print(f"Successfully updated images for product {product_id}")
        except requests.exceptions.HTTPError as err:
            print(f"HTTP error occurred while updating images for product {product_id}: {err}")
            print(f"Response: {response.text}")
        except Exception as err:
            print(f"An error occurred while updating images for product {product_id}: {err}")

    # Function to update a product on Shopify
    def update_product(product_id, updated_data):
        product_id = str(int(product_id)) if not pd.isna(product_id) else None
        if not product_id:
            print("Product ID not provided, creating a new product.")
            return create_new_product(updated_data)

        url = f"{BASE_URL}/products/{product_id}.json"
        updated_data["product"] = clean_data(updated_data["product"])
        response = requests.put(url, headers=headers, json=updated_data)

        if response.status_code == 404:  # Product not found, create it
            print(f"Product {product_id} not found. Creating new product.")
            return create_new_product(updated_data)
        elif response.status_code == 200:
            print(f"Successfully updated product {product_id}")
            variant_id = None
            product_payload = response.json().get("product", {}) if response.content else {}
            variants = product_payload.get("variants", []) if isinstance(product_payload, dict) else []
            if variants:
                variant_id = variants[0].get("id")
            if not variant_id:
                variant_id = (
                    updated_data.get("product", {})
                    .get("variants", [{}])[0]
                    .get("id")
                )
            return product_id, variant_id
        else:
            print(f"Failed to update product {product_id}: {response.status_code}, {response.text}")
            return None, None
        
    def update_product_by_handle(handle, updated_data):
        if not handle:
            print("Handle not provided. Cannot update or create product.")
            return

        # Fetch product by handle
        url = f"{BASE_URL}/products.json?handle={handle}"
        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            products = response.json().get('products', [])
            if products:
                print(f"Product with handle '{handle}' found. Updating product.")
                product_id = products[0]['id']  # Capture the product ID

                # Grab the first variant’s ID if it exists
                variants = updated_data.get("variants", [])
                if variants:
                    variant_id = variants[0].get("id")
                else:
                    variant_id = None

                url_update = f"{BASE_URL}/products/{products[0]['id']}.json"
                updated_data["product"]["handle"] = handle  # Ensure handle is included
                updated_data["product"] = clean_data(updated_data["product"])  # Clean data
                # Clean the data before sending
                updated_data = clean_data(updated_data)
                response_update = requests.put(url_update, headers=headers, json=updated_data)
                if response_update.status_code == 200:
                    print(f"Successfully updated product with handle '{handle}' and ID {product_id}.")
                    if variant_id:
                        print(f"First variant ID: {variant_id}")
                    return product_id, variant_id  # Return the product ID for further use
                else:
                    print(f"Failed to update product with handle '{handle}': {response_update.status_code}, {response_update.text}")
            else:
                print(f"Product with handle '{handle}' not found. Creating new product.")
                updated_data["product"]["handle"] = handle
                updated_data["product"] = clean_data(updated_data["product"])  # Clean data
                print("Updated Data for Product Creation:", json.dumps(updated_data, indent=4))

                # Call create_new_product and capture product_id and variant_id
                product_id, variant_id = create_new_product(updated_data)  # Capture both IDs

                # Log or use the IDs as needed
                if product_id:
                    print(f"Product created successfully with ID {product_id} and handle {handle}")
                    if variant_id:
                        print(f"First variant ID: {variant_id}")
                    else:
                        print("No variants created for this product.")
                    return product_id, variant_id  # Return both IDs

                else:
                    print("Failed to create a new product.")
                

                
        else:
            print(f"Failed to fetch product by handle '{handle}': {response.status_code}, {response.text}")

    def get_all_products_forsku():
        products = []
        url = f"{BASE_URL}/products.json?limit=250&fields=id,title,body_html,handle,tags,vendor,product_type,variants,images,created_at,updated_at,status,published_at,published_scope,template_suffix,options"

        print("Starting to fetch products...")
        page = 1  # Page counter

        while url:
            print(f"Fetching page {page} of products...")
            response = requests.get(url, headers=headers)
            if response.status_code != 200:
                print(f"Error fetching products: {response.status_code}")
                break

            batch = response.json().get('products', [])
            if not batch:  # Check if the current batch is empty to avoid unnecessary processing
                print("No more products found.")
                break

            products.extend(batch)

            # Properly handle pagination if more products are available
            link_header = response.headers.get('Link')
            next_url = None
            if link_header:
                links = link_header.split(',')
                for link in links:
                    if 'rel="next"' in link:
                        next_url = link.split(';')[0].strip('<> ').replace('&amp;', '&')
                        break

            if next_url:
                url = next_url
                page += 1
            else:
                url = None

        print(f"Finished fetching products. Total products fetched: {len(products)}")
        return products


    def update_product_by_sku(sku, updated_data):
        if not sku:
            print("SKU not provided. Cannot update or create product.")
            return None, None
        
        print("Updated Data Received:", json.dumps(updated_data, indent=4))


        # Fetch all products
        products = get_all_products_forsku()
        found_variant = None

        # Check if any existing variant matches the SKU
        for product in products:
            for variant in product['variants']:
                if variant['sku'] == sku:
                    found_variant = variant
                    break
            if found_variant:
                break

        if found_variant:
            product_id = found_variant['product_id']  # Assuming this key exists in your data structure
            variant_id = found_variant['id']

            # Update the product with the product ID
            url_update = f"{BASE_URL}/products/{product_id}.json"
            updated_data["product"] = clean_data(updated_data["product"])  # Clean the data before sending
            response_update = requests.put(url_update, headers=headers, json=updated_data)
            if response_update.status_code == 200:
                updated_fields = ", ".join(updated_data["product"].keys())
                print(f"Successfully updated product with SKU '{sku}' and ID {product_id}. Updated fields: {updated_fields}.")
                return product_id, variant_id
            else:
                print(f"Failed to update product with SKU '{sku}': {response_update.status_code}, {response_update.text}")

        print(f"No product found with SKU '{sku}'. Attempting to create new product.")

        # If no product with the given SKU exists, create a new one
        variant_details = updated_data["product"]["variants"][0] if "variants" in updated_data["product"] and len(updated_data["product"]["variants"]) > 0 else {}
        updated_data["product"]["variants"] = [{
            'sku': sku,
            'price': variant_details.get("price"),
            'weight': variant_details.get("weight"),
            'weight_unit': variant_details.get("weight_unit", "kg")  # assuming default weight unit if not specified
        }]  # Add SKU, price, and weight to the variants in the product data
        updated_data["product"] = clean_data(updated_data["product"])  # Clean the data
        print("Updated Data for Product Creation:", json.dumps(updated_data, indent=4))
        product_id, variant_id = create_new_product(updated_data)
        if product_id:
            print(f"Product created successfully with SKU {sku} and ID {product_id}")
            if variant_id:
                print(f"First variant ID: {variant_id}")
            return product_id, variant_id
        else:
            print("Failed to create a new product.")
            return None, None


    def delete_variant(variant_id):
        if not variant_id:
            print("Variant ID not provided. Cannot delete variant.")
            return
        
        url = f"{BASE_URL}/variants/{variant_id}.json"
        response = requests.delete(url, headers=headers)
        
        if response.status_code in [200, 204]:
            print(f"Successfully deleted variant with ID {variant_id}.")
        else:
            print(f"Failed to delete variant with ID {variant_id}: {response.status_code}, {response.text}")


    def create_new_product(data):
        # Check if the necessary product information is available
        if "handle" not in data["product"] and "sku" not in data["product"].get("variants", [{}])[0]:
            print("Neither handle nor SKU provided in product data. Cannot create product.")
            return None, None  # Return None for both product ID and variant ID

        url = f"{BASE_URL}/products.json"
        response = requests.post(url, headers=headers, json=data)
        if response.status_code in [200, 201]:
            product = response.json().get("product", {})
            product_id = product.get("id")
            variants = product.get("variants", [])  # Get the variants from the response
            variant_id = variants[0].get("id") if variants else None  # Extract the first variant ID if available
            
            if product_id:
                identifier = data['product'].get('handle', data['product'].get("variants", [{}])[0].get("sku"))
                print(f"Product created successfully with ID {product_id} and identifier {identifier}")
                if variant_id:
                    print(f"First Variant ID: {variant_id}")
                return product_id, variant_id  # Return both the product ID and the first variant ID
        elif response.status_code == 422:
            print(f"Validation error while creating product: {response.text}")
        elif response.status_code == 429:
            print("Rate limit exceeded. Pausing to retry...")
            time.sleep(1)  # Delay before retrying
            return create_new_product(data)  # Retry the request
        else:
            print(f"Failed to create product: {response.status_code}, {response.text}")

        return None, None  # Return None for both product ID and variant ID if creation fails


    def update_or_create_variant(product_id, variant_id, updated_data):
        if not product_id:
            print("Product ID is required to create or update a variant.")
            return False, None, None, None

        inventory_qty = None
        if isinstance(updated_data, dict):
            inventory_qty = updated_data.get("variant", {}).get("inventory_quantity")

        if variant_id:
            url = f"{BASE_URL}/variants/{variant_id}.json"
            cleaned_data = clean_data(updated_data)
            cleaned_data["variant"].pop('inventory_quantity', None)

            try:
                response = requests.put(url, headers=headers, json=cleaned_data)
            except requests.RequestException as exc:
                print(f"Network error while updating variant {variant_id}: {exc}")
                return False, None, None, {"error": str(exc)}

            response_status = response.status_code
            try:
                response_json = response.json()
            except ValueError:
                response_json = {"raw": response.text}

            if response_status == 200:
                print(f"Successfully updated variant {variant_id}")
                variant_payload = (
                    response_json.get("variant", {})
                    if isinstance(response_json, dict)
                    else {}
                )

                if inventory_qty is not None:
                    inventory_item_id = variant_payload.get("inventory_item_id")
                    set_inventory_level(inventory_item_id, inventory_qty)

                if isinstance(response_json, dict) and response_json.get("errors"):
                    print(
                        f"Variant update for {variant_id} completed with errors: {response_json.get('errors')}"
                    )

                updated_variant_id = variant_payload.get("id", variant_id)
                return True, updated_variant_id, response_status, response_json

            elif response_status == 404:
                print(f"Variant {variant_id} not found. Creating new variant.")
                created_variant_id = create_new_variant(product_id, updated_data)
                return (created_variant_id is not None), created_variant_id, response_status, response_json

            print(f"Failed to update variant {variant_id}: {response_status}, {response.text}")
            return False, None, response_status, response_json

        print("Variant ID not provided, creating a new variant.")
        created_variant_id = create_new_variant(product_id, updated_data)
        return (created_variant_id is not None), created_variant_id, None, None


    def update_or_create_variant_by_handle(handle, variant_data):
        if not handle:
            print("Handle not provided. Cannot update or create variant.")
            return None

        # Fetch product by handle
        url = f"{BASE_URL}/products.json?handle={handle}"
        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            products = response.json().get('products', [])
            if products:
                product_id = products[0]['id']
                variants_url = f"{BASE_URL}/products/{product_id}/variants.json"
                variants_response = requests.get(variants_url, headers=headers)

                if variants_response.status_code == 200:
                    variants = variants_response.json().get('variants', [])

                    def normalize_option(value):
                        if value is None:
                            return None
                        if isinstance(value, str):
                            value = value.strip()
                            return value if value else None
                        return str(value)

                    target_option1 = normalize_option(variant_data["variant"].get("option1"))
                    target_option2 = normalize_option(variant_data["variant"].get("option2"))
                    target_option3 = normalize_option(variant_data["variant"].get("option3"))

                    def is_matching_variant(variant):
                        if normalize_option(variant.get("option1")) != target_option1:
                            return False
                        if target_option2 is not None and normalize_option(variant.get("option2")) != target_option2:
                            return False
                        if target_option3 is not None and normalize_option(variant.get("option3")) != target_option3:
                            return False
                        return True

                    existing_variant = next((v for v in variants if is_matching_variant(v)), None)
                    if existing_variant:
                        variant_id = existing_variant['id']
                        print(f"Existing variant found with ID {variant_id}. Updating variant.")
                        (
                            success,
                            updated_variant_id,
                            response_status,
                            response_payload,
                        ) = update_or_create_variant(product_id, variant_id, variant_data)

                        if success and updated_variant_id:
                            if isinstance(response_payload, dict):
                                variant_response = response_payload.get("variant", {}) or {}
                                requested_image_id = variant_data["variant"].get("image_id")
                                requested_media_id = variant_data["variant"].get("media_id")
                                response_image_id = variant_response.get("image_id")

                                if (
                                    (requested_image_id or requested_media_id)
                                    and not response_image_id
                                ):
                                    print(
                                        f"Warning: Shopify did not return an image_id for variant {updated_variant_id} "
                                        f"after requesting image assignment (image_id={requested_image_id}, "
                                        f"media_id={requested_media_id})."
                                    )

                            return updated_variant_id  # Return the ID of the updated variant

                        failure_details = []
                        if response_status:
                            failure_details.append(f"status {response_status}")

                        if isinstance(response_payload, dict):
                            if response_payload.get("errors"):
                                failure_details.append(f"errors: {response_payload.get('errors')}")
                            elif response_payload.get("raw"):
                                failure_details.append(f"response: {response_payload.get('raw')}")
                        elif response_payload:
                            failure_details.append(f"response: {response_payload}")

                        requested_image_id = variant_data["variant"].get("image_id")
                        requested_media_id = variant_data["variant"].get("media_id")
                        if requested_image_id or requested_media_id:
                            failure_details.append(
                                "requested "
                                f"image_id={requested_image_id} media_id={requested_media_id}"
                            )

                        detail_message = "; ".join(detail for detail in failure_details if detail)
                        print(
                            f"Variant update failed for handle '{handle}' (variant ID {variant_id}). "
                            + (detail_message if detail_message else "Check the response details above.")
                        )
                        return None
                    else:
                        print("No matching variant found. Creating a new variant.")
                        new_variant_id = create_new_variant(product_id, variant_data)
                        if new_variant_id:
                            print(f"New variant created with ID {new_variant_id}.")
                        return new_variant_id  # Return the ID of the newly created variant
                else:
                    print(f"Failed to fetch variants for product ID {product_id}: {variants_response.status_code}, {variants_response.text}")
            else:
                print(f"Product with handle '{handle}' not found. Cannot create variant.")
        else:
            print(f"Failed to fetch product by handle '{handle}': {response.status_code}, {response.text}")

        return None  # Return None if no variant ID could be retrieved or created



    def create_new_variant(product_id, updated_data):
        url = f"{BASE_URL}/products/{product_id}/variants.json"

        # Ensure required fields are present
        required_options = {
            "option1": updated_data["variant"].get("option1", "Default Option1"),
            "option2": updated_data["variant"].get("option2", "Default Option2"),
            "option3": updated_data["variant"].get("option3", "")
        }
        updated_data["variant"].update(required_options)

        inventory_qty = updated_data["variant"].get("inventory_quantity") if isinstance(updated_data, dict) else None

        # Clean the data to remove NaN or unsupported values
        cleaned_data = clean_data(updated_data)
        cleaned_data["variant"]["product_id"] = product_id  # Explicitly link to product

        response = requests.post(url, headers=headers, json=cleaned_data)
        if response.status_code in [200, 201]:
            # Variant created successfully
            created_variant = response.json().get('variant', {})
            variant_id = created_variant.get('id')
            print(f"Variant created successfully for product ID {product_id}. Variant ID: {variant_id}")

            if inventory_qty is not None:
                inventory_item_id = created_variant.get("inventory_item_id")
                set_inventory_level(inventory_item_id, inventory_qty)

            # Fetch all variants for the product after creating the new variant
            variants_url = f"{BASE_URL}/products/{product_id}/variants.json"
            variants_response = requests.get(variants_url, headers=headers)
            
            if variants_response.status_code == 200:
                variants = variants_response.json().get('variants', [])
                for variant in variants:
                    # Identify and delete the "Default Title" variant
                    if variant.get('title') == "Default Title":
                        print(f"Default Title variant found with ID {variant['id']}. Deleting it...")
                        delete_variant(variant['id'])
                        break  # Exit the loop after deleting the default variant
            else:
                print(f"Failed to fetch variants for product ID {product_id}: {variants_response.status_code}, {variants_response.text}")

            # Return the ID of the newly created variant
            return variant_id

        elif response.status_code == 422:
            error_message = response.json().get("errors", {})
            if "base" in error_message and "already exists" in error_message["base"][0]:
                print("Variant already exists. Skipping creation.")
            else:
                print(f"Failed to create variant: {response.status_code}, {response.text}")
        elif response.status_code == 429:
            print("Rate limit exceeded. Pausing to retry...")
            time.sleep(1)  # Delay for one second before retrying
            return create_new_variant(product_id, updated_data)
        else:
            print(f"Failed to create variant: {response.status_code}, {response.text}")

        # Return None if the creation failed
        return None






    # Function to delete a metafield
    def delete_metafield(owner_id, metafield_id):
        url = f"{BASE_URL}/metafields/{metafield_id}.json"
        response = requests.delete(url, headers=headers)
        if response.status_code == 200:
            print(f"Successfully deleted metafield {metafield_id} for owner {owner_id}")
        else:
            print(f"Failed to delete metafield {metafield_id} for owner {owner_id}: {response.status_code}, {response.text}")

    # Function to create a staged upload
    def staged_upload_create(filename, mime_type):
        query = """
        mutation {
        stagedUploadsCreate(input: {
            resource: FILE,
            filename: "%s",
            mimeType: "%s",
            httpMethod: POST
        }) {
            stagedTargets {
            url
            parameters {
                name
                value
            }
            }
            userErrors {
            field
            message
            }
        }
        }
        """ % (filename, mime_type)

        response = requests.post(GRAPHQL_URL, json={"query": query}, headers=graphql_headers)
        data = response.json()
        if "data" in data and "stagedUploadsCreate" in data["data"]:
            return data["data"]["stagedUploadsCreate"]["stagedTargets"][0]
        else:
            print("Error in staged upload creation:", data)
            return None

    # Function to upload the file to the staging URL
    def upload_file_to_staging(staging_target, file_path):
        url = staging_target["url"]
        files = {"file": open(file_path, "rb")}
        form_data = {param["name"]: param["value"] for param in staging_target["parameters"]}
        response = requests.post(url, data=form_data, files=files)

        # Parsing the XML response to get the file location
        if response.status_code == 201:
            print(f"File {file_path} successfully uploaded to staging URL.")
            xml_response = ET.fromstring(response.text)
            location_url = xml_response.find('Location').text  # Extracting Location URL from XML
            return location_url
        else:
            print(f"Failed to upload file. Status: {response.status_code}, {response.text}")
            return None

    # Function to commit the file with fileCreate
    def commit_file_to_shopify(file_name, original_source):
        query = """
        mutation fileCreate($files: [FileCreateInput!]!) {
        fileCreate(files: $files) {
            files {
            id  # Fetch the gid after file creation
            alt
            createdAt
            ... on GenericFile {
                url
            }
            ... on MediaImage {
                image {
                url
                }
            }
            }
            userErrors {
            code
            field
            message
            }
        }
        }
        """
        variables = {
            "files": [
                {
                    "alt": file_name,
                    "originalSource": original_source
                }
            ]
        }
        response = requests.post(GRAPHQL_URL, json={"query": query, "variables": variables}, headers=graphql_headers)
        data = response.json()
        if "data" in data and "fileCreate" in data["data"]:
            file_info = data["data"]["fileCreate"]["files"]
            gid = file_info[0]["id"]
            print(f"File {file_name} successfully committed to Shopify.")
            print(f"GID: {gid}")  # Print the gid here
            return file_info
        else:
            print("Error in file commit:", data)
            return None

    # Function to get image URL for a given GID
    def get_image_url_for_gid(gid):
        query = f"""
        {{
        node(id: "{gid}") {{
            ... on GenericFile {{
            url
            }}
            ... on MediaImage {{
            image {{
                url
            }}
            }}
        }}
        }}
        """
        response = requests.post(GRAPHQL_URL, json={"query": query}, headers=graphql_headers)
        if response.status_code == 200:
            data = response.json()
            node = (((data or {}).get('data') or {}).get('node') or {})
            url = None
            if 'url' in node:
                url = node['url']
            elif 'image' in node and node['image']:
                url = node['image'].get('url')
            return url
        else:
            print(f"Failed to get image URL for GID {gid}")
            return None

   
        
    # Function to upload a generic file (image, PDF, etc.) to Shopify
    def upload_image_to_shopify(file_path):
        filename = os.path.basename(file_path)

        # Normalize the filename before accessing it
        normalized_filename = normalize_filename(filename)
        folder_path = os.path.dirname(file_path)

        file_path = os.path.join(os.path.dirname(file_path), normalized_filename)


        # Check if file exists after normalization
        if not os.path.exists(file_path):
            print(f"Image file {normalized_filename} not found in local folder.")
            return None, None


          # Ensure the filename is URL-encoded
        encoded_filename = encode_filename(filename)


        # Extract the file name and MIME type
        mime_type = guess_mime_type(filename)

        # Step 1: Create the staged upload
        staging_target = staged_upload_create(filename, mime_type)
        if not staging_target:
            return None, None

        # Step 2: Upload the file to the staging URL
        location_url = upload_file_to_staging(staging_target, file_path)
        if not location_url:
            return None, None

        # Step 3: Commit the file to Shopify
        file_info = commit_file_to_shopify(filename, location_url)
        if file_info:
            # file_info contains the files array
            file_data = file_info[0]
            gid = file_data['id']
            # Get the URL
            url = None
            if 'url' in file_data:
                url = file_data['url']
            elif 'image' in file_data and file_data['image']:
                url = file_data['image'].get('url')
            if not url:
                # Get URL via get_image_url_for_gid
                url = get_image_url_for_gid(gid)
                if not url:
                    print(f"⚠️ Unable to resolve URL for uploaded file with GID {gid}")

            return url, gid
        else:
            return None, None

    # Function to get all files from Shopify
    def get_all_files():
        all_files = {}
        has_next_page = True
        cursor = None

        while has_next_page:
            query = f"""
            {{
            files(first: 250{' , after: "' + cursor + '"' if cursor else ''}) {{
                edges {{
                node {{
                    id
                    alt
                    ... on GenericFile {{
                    url
                    }}
                    ... on MediaImage {{
                    image {{
                        url
                    }}
                    }}
                }}
                cursor
                }}
                pageInfo {{
                hasNextPage
                }}
            }}
            }}
            """
            response = requests.post(GRAPHQL_URL, json={"query": query}, headers=graphql_headers)
            if response.status_code == 200:
                data = response.json()
                if "data" in data and "files" in data["data"]:
                    for file in data["data"]["files"]["edges"]:
                        node = file["node"]
                        gid = node["id"]
                        alt = node["alt"]
                        url = None
                        if 'url' in node:
                            url = node['url']
                        elif 'image' in node and node['image']:
                            url = node['image'].get('url')
                        if alt:
                            remember_file_reference(all_files, alt, gid, url)
                        filename = extract_filename_from_value(url)
                        if filename:
                            remember_file_reference(all_files, filename, gid, url)
                        cursor = file["cursor"]

                    has_next_page = data["data"]["files"]["pageInfo"]["hasNextPage"]
                else:
                    print("No files found or error in response.")
                    return None
            else:
                print(f"Error fetching files. Status code: {response.status_code}")
                return None

        return all_files

    # Function to update or create metafields for a product
    def upsert_metafield(owner_type, owner_id, namespace, key, metafield_data, current_metafields_dict):
        """Create or update a metafield for the given owner."""
        metafield_key = f"{namespace}.{key}"
        existing_metafield_id = current_metafields_dict.get(metafield_key)

        if existing_metafield_id:
            payload = {"metafield": {"id": existing_metafield_id}}
            payload["metafield"].update(metafield_data["metafield"])
            url = f"{BASE_URL}/metafields/{existing_metafield_id}.json"
            response = requests.put(url, headers=headers, json=payload)
            action = "updated"
        else:
            endpoint = "products" if owner_type == "product" else "variants"
            url = f"{BASE_URL}/{endpoint}/{owner_id}/metafields.json"
            response = requests.post(url, headers=headers, json=metafield_data)
            action = "created"

        if response.status_code in [200, 201]:
            metafield_response = response.json().get("metafield", {})
            metafield_id = metafield_response.get("id", existing_metafield_id)
            if metafield_id:
                current_metafields_dict[metafield_key] = metafield_id
            print(
                f"✅ Successfully {action} metafield {namespace}.{key} for {owner_type} {owner_id}"
            )
        else:
            print(
                f"❌ Failed to {action} metafield {namespace}.{key} for {owner_type} {owner_id}: {response.status_code}"
            )
            print(f"⚠️ Response Body: {response.text}")

    def update_metafields(handle, metafields, existing_files, row_index, df):
        product_id = handle
        if not product_id:
            print(f"Skipping metafield update for missing product ID.")
            return

        current_metafields_url = f"{BASE_URL}/products/{product_id}/metafields.json"
        response = requests.get(current_metafields_url, headers=headers)
        current_metafields = response.json().get('metafields', []) if response.status_code == 200 else []
        current_metafields_dict = {f"{mf['namespace']}.{mf['key']}": mf['id'] for mf in current_metafields}

        for column, value in metafields.items():
            key_type_str = column.replace('Metafield: ', '').split(' ')
            key = key_type_str[0]
            field_type = key_type_str[1].replace('[', '').replace(']', '') if len(key_type_str) > 1 else 'single_line_text_field'

            namespace, key = key.split('.')

            if pd.isna(value) or value is None:
                metafield_key = f"{namespace}.{key}"
                if metafield_key in current_metafields_dict:
                    delete_metafield(product_id, current_metafields_dict[metafield_key])
                    current_metafields_dict.pop(metafield_key, None)
                continue

            metafield_data = None

            if field_type == 'file_reference':
                value_gid = None
                filename = None
                if isinstance(value, str):
                    candidate = value.strip()
                    if candidate.startswith('gid://'):
                        value_gid = candidate
                    elif candidate.startswith('http'):
                        filename = extract_filename_from_value(candidate)
                        existing_entry = fetch_file_reference(existing_files, filename)
                        value_gid = existing_entry[0] if existing_entry else None
                    else:
                        filename = extract_filename_from_value(candidate)
                        existing_entry = fetch_file_reference(existing_files, filename)
                        if existing_entry:
                            value_gid = existing_entry[0]
                            set_dataframe_cell(df, row_index, column, existing_entry[1] or value_gid)
                        else:
                            file_path_local = resolve_local_asset_path(candidate)
                            if file_path_local:
                                url, gid = upload_image_to_shopify(file_path_local)
                                if gid:
                                    remember_file_reference(existing_files, filename or candidate, gid, url)
                                    value_gid = gid
                                    if url:
                                        set_dataframe_cell(df, row_index, column, url)
                                    else:
                                        set_dataframe_cell(df, row_index, column, value_gid)
                                else:
                                    print(f"❌ Failed to upload file {candidate} for metafield {namespace}.{key}")
                                    continue
                            else:
                                print(f"⚠️ File {candidate} not found locally for metafield {namespace}.{key}")
                                continue
                else:
                    print(f"⚠️ Skipping non-string value for metafield {namespace}.{key}")
                    continue

                if value_gid:
                    metafield_data = {
                        "metafield": {
                            "namespace": namespace,
                            "key": key,
                            "value": value_gid,
                            "type": field_type.strip()
                        }
                    }
                else:
                    print(f"❌ Cannot determine GID for metafield file value '{value}'")
                    continue

            elif field_type == 'list.file_reference':
                if isinstance(value, str):
                    file_names = [filename.strip() for filename in value.split(',') if filename.strip()]
                    file_gids = []

                    print(f"🔍 Processing metafield '{key}' for product {product_id} with {len(file_names)} files: {file_names}")

                    for raw_name in file_names:
                        value_gid = None
                        filename = extract_filename_from_value(raw_name)

                        if raw_name.startswith('gid://'):
                            value_gid = raw_name
                        elif raw_name.startswith('http'):
                            existing_entry = fetch_file_reference(existing_files, filename)
                            value_gid = existing_entry[0] if existing_entry else None
                        else:
                            existing_entry = fetch_file_reference(existing_files, filename)
                            if existing_entry:
                                value_gid = existing_entry[0]
                                set_dataframe_cell(df, row_index, column, existing_entry[1] or value_gid)
                            else:
                                file_path_local = resolve_local_asset_path(raw_name)
                                if file_path_local:
                                    print(f"📤 Uploading {raw_name} to Shopify...")
                                    url, gid = upload_image_to_shopify(file_path_local)
                                    if gid:
                                        remember_file_reference(existing_files, filename or raw_name, gid, url)
                                        value_gid = gid
                                        if url:
                                            set_dataframe_cell(df, row_index, column, url)
                                        else:
                                            set_dataframe_cell(df, row_index, column, value_gid)
                                    else:
                                        print(f"❌ Failed to upload file {raw_name}")
                                        continue
                                else:
                                    print(f"⚠️ File {raw_name} not found locally.")
                                    continue

                        if value_gid:
                            file_gids.append(value_gid)
                        else:
                            print(f"❌ Cannot find GID for file {raw_name}")

                    print(f"📝 Final GID list for metafield '{key}': {file_gids}")

                    if file_gids:
                        metafield_data = {
                            "metafield": {
                                "namespace": namespace,
                                "key": key,
                                "value": json.dumps(file_gids),
                                "type": 'list.file_reference'
                            }
                        }
                        print(f"📡 Sending metafield update to Shopify: {metafield_data}")
                    else:
                        print(f"❌ Skipping metafield update for '{key}' because the file list is empty (avoiding 422 error).")
                        continue
                else:
                    print(f"⚠️ Skipping non-string value for metafield {key}.")
                    continue

            elif field_type == 'metaobject_reference':
                value_gid = resolve_metaobject_reference_value(
                    "PRODUCT", namespace, key, value
                )
                if not value_gid:
                    continue

                metafield_data = {
                    "metafield": {
                        "namespace": namespace,
                        "key": key,
                        "value": value_gid,
                        "type": field_type.strip()
                    }
                }

            elif field_type == 'list.metaobject_reference':
                value_gids = resolve_metaobject_reference_list(
                    "PRODUCT", namespace, key, value
                )
                if not value_gids:
                    print(
                        f"❌ Skipping metafield update for {namespace}.{key} because no "
                        "metaobject references could be resolved."
                    )
                    continue

                metafield_data = {
                    "metafield": {
                        "namespace": namespace,
                        "key": key,
                        "value": json.dumps(value_gids),
                        "type": field_type.strip()
                    }
                }

            elif field_type == 'url':
                if isinstance(value, str) and value.strip():
                    raw_value = value.strip()
                    final_url = None
                    filename = extract_filename_from_value(raw_value)

                    if raw_value.startswith('http'):
                        existing_entry = fetch_file_reference(existing_files, filename)
                        final_url = existing_entry[1] if existing_entry and existing_entry[1] else raw_value
                    else:
                        existing_entry = fetch_file_reference(existing_files, filename)
                        if existing_entry and existing_entry[1]:
                            final_url = existing_entry[1]
                        else:
                            file_path_local = resolve_local_asset_path(raw_value)
                            if not file_path_local:
                                print(f"⚠️ File {raw_value} not found for URL metafield {namespace}.{key}")
                                continue
                            url, gid = upload_image_to_shopify(file_path_local)
                            if url:
                                remember_file_reference(existing_files, filename or raw_value, gid, url)
                                final_url = url
                            else:
                                print(f"❌ Failed to upload file {raw_value} for URL metafield {namespace}.{key}")
                                continue

                    if not final_url:
                        print(f"⚠️ Skipping URL metafield {namespace}.{key} due to missing URL value.")
                        continue

                    set_dataframe_cell(df, row_index, column, final_url)

                    metafield_data = {
                        "metafield": {
                            "namespace": namespace,
                            "key": key,
                            "value": final_url,
                            "type": 'url'
                        }
                    }
                else:
                    print(f"⚠️ Skipping invalid URL metafield value for {namespace}.{key}")
                    continue

            else:
                if field_type == 'rich_text_field':
                    try:
                        value = json.dumps(value) if isinstance(value, dict) else value
                    except Exception as e:
                        print(f"Error serializing JSON for {namespace}.{key}: {e}")
                        continue
                elif field_type in ('single_line_text_field', 'multi_line_text_field'):
                    value = format_metafield_text_value(value)

                metafield_data = {
                    "metafield": {
                        "namespace": namespace,
                        "key": key,
                        "value": value,
                        "type": field_type.strip()
                    }
                }
                print(f"Other metafield data: {metafield_data}")

            if metafield_data:
                upsert_metafield(
                    "product", product_id, namespace, key, metafield_data, current_metafields_dict
                )

    def update_variant_metafields(variant_id, metafields, existing_files, row_index, df):
        if not variant_id:
            print("Skipping metafield update for missing variant ID.")
            return

        current_metafields_url = f"{BASE_URL}/variants/{variant_id}/metafields.json"
        response = requests.get(current_metafields_url, headers=headers)
        current_metafields = response.json().get('metafields', []) if response.status_code == 200 else []
        current_metafields_dict = {f"{mf['namespace']}.{mf['key']}": mf['id'] for mf in current_metafields}

        for column, value in metafields.items():
            key_type_str = column.replace('Variant Metafield: ', '').split(' ')
            key = key_type_str[0]
            field_type = key_type_str[1].replace('[', '').replace(']', '') if len(key_type_str) > 1 else 'single_line_text_field'

            namespace, key = key.split('.')

            if pd.isna(value) or value is None:
                metafield_key = f"{namespace}.{key}"
                if metafield_key in current_metafields_dict:
                    delete_metafield(variant_id, current_metafields_dict[metafield_key])
                    current_metafields_dict.pop(metafield_key, None)
                continue

            metafield_data = None

            if field_type == 'file_reference':
                value_gid = None
                filename = None
                if isinstance(value, str):
                    candidate = value.strip()
                    if candidate.startswith('gid://'):
                        value_gid = candidate
                    elif candidate.startswith('http'):
                        filename = extract_filename_from_value(candidate)
                        existing_entry = fetch_file_reference(existing_files, filename)
                        value_gid = existing_entry[0] if existing_entry else None
                    else:
                        filename = extract_filename_from_value(candidate)
                        existing_entry = fetch_file_reference(existing_files, filename)
                        if existing_entry:
                            value_gid = existing_entry[0]
                            set_dataframe_cell(df, row_index, column, existing_entry[1] or value_gid)
                        else:
                            file_path_local = resolve_local_asset_path(candidate)
                            if file_path_local:
                                url, gid = upload_image_to_shopify(file_path_local)
                                if gid:
                                    remember_file_reference(existing_files, filename or candidate, gid, url)
                                    value_gid = gid
                                    if url:
                                        set_dataframe_cell(df, row_index, column, url)
                                    else:
                                        set_dataframe_cell(df, row_index, column, value_gid)
                                else:
                                    print(f"❌ Failed to upload file {candidate} for variant metafield {namespace}.{key}")
                                    continue
                            else:
                                print(f"⚠️ File {candidate} not found locally for variant metafield {namespace}.{key}")
                                continue
                else:
                    continue

                if value_gid:
                    metafield_data = {
                        "metafield": {
                            "namespace": namespace,
                            "key": key,
                            "value": value_gid,
                            "type": field_type.strip()
                        }
                    }
                else:
                    print(f"❌ Cannot determine GID for variant metafield file value '{value}'")
                    continue

            elif field_type == 'list.file_reference':
                if isinstance(value, str):
                    file_names = [filename.strip() for filename in value.split(',') if filename.strip()]
                    file_gids = []

                    for raw_name in file_names:
                        value_gid = None
                        filename = extract_filename_from_value(raw_name)

                        if raw_name.startswith('gid://'):
                            value_gid = raw_name
                        elif raw_name.startswith('http'):
                            existing_entry = fetch_file_reference(existing_files, filename)
                            value_gid = existing_entry[0] if existing_entry else None
                        else:
                            existing_entry = fetch_file_reference(existing_files, filename)
                            if existing_entry:
                                value_gid = existing_entry[0]
                                set_dataframe_cell(df, row_index, column, existing_entry[1] or value_gid)
                            else:
                                file_path_local = resolve_local_asset_path(raw_name)
                                if file_path_local:
                                    url, gid = upload_image_to_shopify(file_path_local)
                                    if gid:
                                        remember_file_reference(existing_files, filename or raw_name, gid, url)
                                        value_gid = gid
                                        if url:
                                            set_dataframe_cell(df, row_index, column, url)
                                        else:
                                            set_dataframe_cell(df, row_index, column, value_gid)
                                    else:
                                        print(f"❌ Failed to upload file {raw_name}")
                                        continue
                                else:
                                    print(f"⚠️ File {raw_name} not found locally.")
                                    continue

                        if value_gid:
                            file_gids.append(value_gid)
                        else:
                            print(f"❌ Cannot find GID for file {raw_name}")

                    if file_gids:
                        metafield_data = {
                            "metafield": {
                                "namespace": namespace,
                                "key": key,
                                "value": json.dumps(file_gids),
                                "type": 'list.file_reference'
                            }
                        }
                    else:
                        print(f"Skipping metafield update for '{key}' because the file list is empty.")
                        continue
                else:
                    print(f"Skipping non-string value for metafield {key}.")
                    continue

            elif field_type == 'metaobject_reference':
                value_gid = resolve_metaobject_reference_value(
                    "PRODUCTVARIANT", namespace, key, value
                )
                if not value_gid:
                    continue

                metafield_data = {
                    "metafield": {
                        "namespace": namespace,
                        "key": key,
                        "value": value_gid,
                        "type": field_type.strip()
                    }
                }

            elif field_type == 'list.metaobject_reference':
                value_gids = resolve_metaobject_reference_list(
                    "PRODUCTVARIANT", namespace, key, value
                )
                if not value_gids:
                    print(
                        f"❌ Skipping variant metafield update for {namespace}.{key} "
                        "because no metaobject references could be resolved."
                    )
                    continue

                metafield_data = {
                    "metafield": {
                        "namespace": namespace,
                        "key": key,
                        "value": json.dumps(value_gids),
                        "type": field_type.strip()
                    }
                }

            elif field_type == 'url':
                if isinstance(value, str) and value.strip():
                    raw_value = value.strip()
                    final_url = None
                    filename = extract_filename_from_value(raw_value)

                    if raw_value.startswith('http'):
                        existing_entry = fetch_file_reference(existing_files, filename)
                        final_url = existing_entry[1] if existing_entry and existing_entry[1] else raw_value
                    else:
                        existing_entry = fetch_file_reference(existing_files, filename)
                        if existing_entry and existing_entry[1]:
                            final_url = existing_entry[1]
                        else:
                            file_path_local = resolve_local_asset_path(raw_value)
                            if not file_path_local:
                                print(f"⚠️ File {raw_value} not found for URL metafield {namespace}.{key}")
                                continue
                            url, gid = upload_image_to_shopify(file_path_local)
                            if url:
                                remember_file_reference(existing_files, filename or raw_value, gid, url)
                                final_url = url
                            else:
                                print(f"❌ Failed to upload file {raw_value} for URL metafield {namespace}.{key}")
                                continue

                    if not final_url:
                        print(f"⚠️ Skipping URL metafield {namespace}.{key} due to missing URL value.")
                        continue

                    set_dataframe_cell(df, row_index, column, final_url)

                    metafield_data = {
                        "metafield": {
                            "namespace": namespace,
                            "key": key,
                            "value": final_url,
                            "type": 'url'
                        }
                    }
                else:
                    print(f"⚠️ Skipping invalid URL metafield value for {namespace}.{key}")
                    continue

            else:
                if field_type == 'rich_text_field':
                    try:
                        value = json.dumps(value) if isinstance(value, dict) else value
                    except Exception as e:
                        print(f"Error serializing JSON for {namespace}.{key}: {e}")
                        continue
                elif field_type in ('single_line_text_field', 'multi_line_text_field'):
                    value = format_metafield_text_value(value)

                metafield_data = {
                    "metafield": {
                        "namespace": namespace,
                        "key": key,
                        "value": value,
                        "type": field_type.strip()
                    }
                }

            if metafield_data:
                upsert_metafield(
                    "variant", variant_id, namespace, key, metafield_data, current_metafields_dict
                )

    def upload_market_prices(price_list_id, market_prices):
        """
        Upload market-specific prices for product variants in bulk.
        """
        url = f"{GRAPHQL_URL}"
        mutation = """
        mutation priceListFixedPricesAdd(
          $priceListId: ID!,
          $prices: [PriceListPriceInput!]!
        ) {
          priceListFixedPricesAdd(priceListId: $priceListId, prices: $prices) {
            prices {
              variant {
                id
              }
            }
            userErrors {
              field
              message
            }
          }
        }
        """
        variables = {
            "priceListId": price_list_id,
            "prices": market_prices
        }

        response = requests.post(url, headers=graphql_headers, json={"query": mutation, "variables": variables})
        if response.status_code == 200:
            data = response.json()
            if "errors" in data:
                print(f"Errors encountered while uploading market prices: {data['errors']}")
            else:
                print(f"Successfully uploaded market prices for {len(market_prices)} variants.")
        else:
            print(f"Failed to upload market prices: {response.status_code}, {response.text}")


    # Function to upload changes from an edited spreadsheet
    def upload_changes_from_spreadsheet(file_path):
        print(f"Reading spreadsheet from: {file_path}")

        df = pd.read_excel(file_path)
        df = df.where(pd.notnull(df), None)

        def has_cell_value(value):
            if value is None:
                return False
            try:
                if pd.isna(value):
                    return False
            except TypeError:
                pass
            if isinstance(value, str):
                return bool(value.strip())
            return True

        def get_cell_value(row_data, column_name, default=None):
            value = row_data.get(column_name, default)
            return value if has_cell_value(value) else default

        def normalize_option_value(value, default=None):
            if not has_cell_value(value):
                return default
            if isinstance(value, str):
                return value.strip()
            if isinstance(value, numbers.Number):
                numeric_value = float(value)
                if numeric_value.is_integer():
                    return str(int(numeric_value))
                return ("{0:f}".format(numeric_value)).rstrip("0").rstrip(".")
            return str(value).strip()

       

        # Get existing files mapping
        print("Fetching all existing files from Shopify...")
        existing_files = get_all_files()
        if existing_files is None:
            existing_files = {}  # To avoid errors
        print(f"Fetched {len(existing_files)} existing files from Shopify.")

        product_images_cache = {}
        product_image_identifier_cache = {}

        def update_product_image_identifiers(product_id_key, images):
            by_src = {}
            by_id = {}
            by_gid = {}

            for image in images or []:
                image_id = image.get("id")
                media_id = image.get("admin_graphql_api_id")
                src = image.get("src")

                entry = {
                    "image_id": image_id,
                    "media_id": media_id,
                    "src": src,
                }

                if src:
                    by_src[src] = entry

                if image_id is not None and not pd.isna(image_id):
                    try:
                        normalized_id = str(int(float(image_id)))
                    except (TypeError, ValueError):
                        normalized_id = str(image_id)
                    by_id[normalized_id] = entry

                if media_id:
                    by_gid[media_id] = entry

            product_image_identifier_cache[product_id_key] = {
                "by_src": by_src,
                "by_id": by_id,
                "by_gid": by_gid,
            }

        def get_product_images(product_id, force_refresh=False):
            if not product_id:
                return []

            product_id_key = str(product_id)

            if force_refresh:
                product_images_cache.pop(product_id_key, None)
                product_image_identifier_cache.pop(product_id_key, None)

            if not force_refresh and product_id_key in product_images_cache:
                return product_images_cache[product_id_key]

            url = f"{BASE_URL}/products/{product_id_key}/images.json"
            try:
                response = requests.get(url, headers=headers)
            except requests.RequestException as exc:
                print(f"Failed to retrieve images for product {product_id_key} due to network error: {exc}")
                product_images_cache[product_id_key] = []
                product_image_identifier_cache[product_id_key] = {
                    "by_src": {},
                    "by_id": {},
                    "by_gid": {},
                }
                return []

            if response.status_code == 200:
                product_images = response.json().get('images', [])
                product_images_cache[product_id_key] = product_images
                update_product_image_identifiers(product_id_key, product_images)
                return product_images

            print(f"Failed to retrieve images for product {product_id_key}: {response.status_code}, {response.text}")
            product_images_cache[product_id_key] = []
            product_image_identifier_cache[product_id_key] = {
                "by_src": {},
                "by_id": {},
                "by_gid": {},
            }
            return []

        def ensure_variant_image(product_id, image_value, alt_text, row_index, handle=None):
            if not image_value:
                return None, None

            def extract_image_identifiers(image_payload):
                if not isinstance(image_payload, dict):
                    return None, None
                return (
                    image_payload.get("id"),
                    image_payload.get("admin_graphql_api_id"),
                )

            resolved_product_id = product_id
            if resolved_product_id:
                resolved_product_id = str(resolved_product_id)
            elif handle:
                lookup_url = f"{BASE_URL}/products.json?handle={handle}"
                try:
                    response = requests.get(lookup_url, headers=headers)
                except requests.RequestException as exc:
                    print(f"Failed to resolve product by handle '{handle}' for variant image: {exc}")
                    return None, None

                if response.status_code == 200:
                    products = response.json().get('products', [])
                    if products:
                        resolved_product_id = str(products[0].get('id'))
                else:
                    print(
                        f"Failed to resolve product by handle '{handle}' for variant image: "
                        f"{response.status_code}, {response.text}"
                    )
                    return None, None

            if not resolved_product_id:
                return None, None

            if isinstance(image_value, str):
                image_reference = image_value.strip()
            else:
                image_reference = str(image_value)

            if not image_reference:
                return None, None

            def update_variant_cell(resolved_url):
                if row_index is not None and resolved_url:
                    set_dataframe_cell(df, row_index, 'Variant Image', resolved_url)

            identifier_lookup = None
            product_images = None

            def load_image_metadata(force=False):
                nonlocal identifier_lookup, product_images
                if force:
                    product_images = get_product_images(resolved_product_id, force_refresh=True)
                elif product_images is None:
                    product_images = get_product_images(resolved_product_id)
                identifier_lookup = product_image_identifier_cache.get(resolved_product_id, {})

            # If an explicit numeric image ID is provided, attempt to use it directly
            try:
                numeric_image_id = int(float(image_reference))
                load_image_metadata()
                lookup_by_id = (identifier_lookup or {}).get("by_id", {})
                numeric_key = str(numeric_image_id)
                entry = lookup_by_id.get(numeric_key)
                if not entry and str(image_reference).strip():
                    entry = lookup_by_id.get(str(image_reference).strip())
                if entry:
                    update_variant_cell(entry.get("src"))
                    return entry.get("image_id"), entry.get("media_id")

                if product_images:
                    for product_image in product_images:
                        try:
                            if int(float(product_image.get('id', 0))) == numeric_image_id:
                                update_variant_cell(product_image.get('src'))
                                return extract_image_identifiers(product_image)
                        except (TypeError, ValueError):
                            continue
            except (ValueError, TypeError):
                pass

            if is_valid_gid(image_reference):
                load_image_metadata()
                lookup_by_gid = (identifier_lookup or {}).get("by_gid", {})
                entry = lookup_by_gid.get(image_reference)
                if entry:
                    update_variant_cell(entry.get("src"))
                    return entry.get("image_id"), entry.get("media_id")

                if product_images:
                    for product_image in product_images:
                        if product_image.get('admin_graphql_api_id') == image_reference:
                            update_variant_cell(product_image.get('src'))
                            return extract_image_identifiers(product_image)

            image_url = image_reference

            if isinstance(image_url, str) and not image_url.startswith(('http://', 'https://', 'gid://')):
                filename = image_url
                existing_entry = fetch_file_reference(existing_files, filename)
                if existing_entry:
                    _, image_url = existing_entry
                else:
                    file_path_local = os.path.join(IMAGE_FOLDER, filename)
                    if not os.path.exists(file_path_local):
                        print(
                            f"Variant image file '{filename}' not found in local folder for product {resolved_product_id} "
                            f"(row {row_index})."
                        )
                        return None, None

                    try:
                        with open(file_path_local, "rb") as image_file:
                            encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
                    except OSError as exc:
                        print(f"Failed to read variant image file '{filename}': {exc}")
                        return None, None

                    image_payload = {
                        "image": {
                            "attachment": encoded_string,
                            "filename": filename,
                        }
                    }
                    if alt_text:
                        image_payload["image"]["alt"] = alt_text

                    try:
                        response = requests.post(
                            f"{BASE_URL}/products/{resolved_product_id}/images.json",
                            headers=headers,
                            json=image_payload,
                        )
                    except requests.RequestException as exc:
                        print(f"Failed to upload variant image '{filename}' for product {resolved_product_id}: {exc}")
                        return None, None

                    if response.status_code in (200, 201):
                        image = response.json().get('image', {})
                        image_url = image.get('src')
                        remember_file_reference(
                            existing_files,
                            filename,
                            image.get('admin_graphql_api_id'),
                            image_url,
                        )
                        product_images_cache.pop(resolved_product_id, None)
                        product_image_identifier_cache.pop(resolved_product_id, None)
                        product_images = None
                        identifier_lookup = None
                        update_variant_cell(image_url)
                    else:
                        print(
                            f"Failed to upload variant image '{filename}' for product {resolved_product_id}: "
                            f"{response.status_code}, {response.text}"
                        )
                        return None, None

            # Check if the product already has this image URL associated
            load_image_metadata()
            lookup_by_src = (identifier_lookup or {}).get("by_src", {})
            src_entry = lookup_by_src.get(image_url)
            if src_entry:
                resolved_src = src_entry.get("src") or image_url
                update_variant_cell(resolved_src)
                return src_entry.get("image_id"), src_entry.get("media_id")

            if product_images:
                for product_image in product_images:
                    if product_image.get('src') == image_url:
                        update_variant_cell(product_image.get('src'))
                        return extract_image_identifiers(product_image)

            # Attach the remote image URL to the product if it isn't already present
            image_payload = {"image": {"src": image_url}}
            if alt_text:
                image_payload["image"]["alt"] = alt_text

            try:
                response = requests.post(
                    f"{BASE_URL}/products/{resolved_product_id}/images.json",
                    headers=headers,
                    json=image_payload,
                )
            except requests.RequestException as exc:
                print(
                    f"Failed to attach variant image from URL '{image_url}' for product {resolved_product_id}: {exc}"
                )
                return None, None

            if response.status_code in (200, 201):
                image = response.json().get('image', {})
                resolved_url = image.get('src')
                if resolved_url:
                    update_variant_cell(resolved_url)
                product_images_cache.pop(resolved_product_id, None)
                product_image_identifier_cache.pop(resolved_product_id, None)
                product_images = None
                identifier_lookup = None
                refreshed_images = get_product_images(resolved_product_id, force_refresh=True)
                product_images = refreshed_images
                identifier_lookup = product_image_identifier_cache.get(resolved_product_id, {})
                for product_image in refreshed_images:
                    if product_image.get('id') == image.get('id') or product_image.get('src') == resolved_url:
                        return extract_image_identifiers(product_image)
                return extract_image_identifiers(image)

            print(
                f"Failed to attach variant image '{image_url}' to product {resolved_product_id}: "
                f"{response.status_code}, {response.text}"
            )
            return None, None

        # Fetch all market names dynamically
        def get_all_market_names():
            query = """
            query Catalogs {
                catalogs(first: 10, type: MARKET) {
                    nodes {
                        ... on MarketCatalog {
                            markets(first: 10) {
                                nodes {
                                    id
                                    name
                                }
                            }
                        }
                    }
                }
            }
            """
            response = requests.post(GRAPHQL_URL, json={"query": query}, headers=headers)
            if response.status_code == 200:
                data = response.json()
                catalogs = ((((data or {}).get("data") or {}).get("catalogs")) or {}).get("nodes", [])
                market_names = []
                for catalog in catalogs:
                    markets = (catalog.get("markets") or {}).get("nodes", [])
                    market_names.extend([market.get("name") for market in markets if market.get("name")])
                return market_names
            print("Failed to fetch markets.")
            return []

        market_names = get_all_market_names()
        print(f"Markets found: {market_names}")

        # Identify market-specific pricing columns
        pricing_columns = {
            column: column.replace("Variant Price / ", "").strip()
            for column in df.columns
            if column.startswith("Variant Price / ")
        }
        print(f"Pricing columns identified: {pricing_columns}")

        def collect_metafields_from_row(row_data, prefix):
            collected = {}
            for col_name in df.columns:
                if not col_name.startswith(prefix):
                    continue

                value = row_data.get(col_name)
                if pd.isna(value):
                    continue

                if (
                    "[rich_text_field]" in col_name
                    and isinstance(value, str)
                    and "<" in value
                    and ">" in value
                ):
                    try:
                        collected[col_name] = html_to_shopify_json(value)
                    except Exception as err:
                        print(f"Error parsing HTML for {col_name}: {err}")
                else:
                    collected[col_name] = value

            return collected

        def get_price_list_id_for_market(market_name):
            query = f"""
            query {{
                catalogs(first: 10, type: MARKET) {{
                    nodes {{
                        ... on MarketCatalog {{
                            markets(first: 10) {{
                                nodes {{
                                    id
                                    name
                                }}
                            }}
                            priceList {{
                                id
                            }}
                        }}
                    }}
                }}
            }}
            """
            response = requests.post(GRAPHQL_URL, json={"query": query}, headers=headers)
            if response.status_code == 200:
                data = response.json()
                catalogs = ((((data or {}).get("data") or {}).get("catalogs")) or {}).get("nodes", [])
                for catalog in catalogs:
                    markets = (catalog.get("markets") or {}).get("nodes", [])
                    if any(market.get("name") == market_name for market in markets):
                        price_list = catalog.get("priceList")
                        if price_list and price_list.get("id"):
                            return price_list.get("id")
            print(f"Price list ID not found for market '{market_name}'.")
            return None

        # Add helper function to add fixed prices for a market
        def add_fixed_price_for_market(price_list_id, variant_id, price_amount):
            mutation = """
            mutation priceListFixedPricesAdd($priceListId: ID!, $prices: [PriceListPriceInput!]!) {
                priceListFixedPricesAdd(priceListId: $priceListId, prices: $prices) {
                    prices {
                        variant {
                            id
                        }
                    }
                    userErrors {
                        field
                        message
                    }
                }
            }
            """
            variables = {
                "priceListId": price_list_id,
                "prices": [
                    {
                        "variantId": f"gid://shopify/ProductVariant/{variant_id}",
                        "price": {
                            "amount": str(price_amount),
                            "currencyCode": "EUR"
                        }
                    }
                ]
            }
            response = requests.post(GRAPHQL_URL, json={"query": mutation, "variables": variables}, headers=headers)
            if response.status_code == 200:
                print(f"Fixed price added for variant ID '{variant_id}' with price '{price_amount}'.")
            else:
                print(f"Failed to add fixed price for variant ID '{variant_id}': {response.text}")
        



        # Collect image filenames from the spreadsheet
        for index, row in df.iterrows():
            # Process product images
            for i in range(1, 21):  # Assuming a maximum of 20 images per product
                image_column = f"Image {i}"
                if image_column in row and row[image_column]:
                    image_value = row[image_column]
                    if isinstance(image_value, str) and not image_value.startswith(('http', 'gid://')):
                        filename = image_value
                        existing_entry = fetch_file_reference(existing_files, filename)
                        if existing_entry:
                            gid, url = existing_entry
                            # Replace cell value with URL
                            set_dataframe_cell(df, index, image_column, url)
                        else:
                            # Upload image to Shopify
                            file_path_local = os.path.join(IMAGE_FOLDER, filename)
                            if os.path.exists(file_path_local):
                                with open(file_path_local, "rb") as image_file:
                                    encoded_string = base64.b64encode(image_file.read()).decode('utf-8')

                                 # Retrieve alt_text or set to default
                                alt_text = row.get("Alt Text", None)  # Use "Alt Text" or the correct column name
                                if not pd.notna(alt_text):  # If alt_text is not available or NaN
                                    alt_text = filename
                                    
                                image_data = {
                                    "image": {
                                        "attachment": encoded_string,
                                        "filename": filename,
                                        "alt": alt_text if pd.notna(alt_text) else filename
                                    }
                                }
                                product_id = row.get('ID')
                                if pd.notna(product_id):
                                    if pd.isna(product_id):
                                        product_id = None
                                    else:
                                        # Convert to string without the .0
                                        product_id = str(int(product_id))

                                    url_api = f"{BASE_URL}/products/{product_id}/images.json"
                                    response = requests.post(url_api, headers=headers, json=image_data)
                                    if response.status_code in [200, 201]:
                                        image = response.json().get('image', {})
                                        image_url = image.get('src')
                                        image_gid = image.get('admin_graphql_api_id')

                                        if not is_valid_gid(image_gid):
                                            uploaded_url, uploaded_gid = upload_image_to_shopify(file_path_local)
                                            if is_valid_gid(uploaded_gid):
                                                image_gid = uploaded_gid
                                                if uploaded_url:
                                                    image_url = uploaded_url

                                        print(f"Successfully uploaded image {filename} to product {product_id}")
                                        # Update cell value with image URL
                                        set_dataframe_cell(df, index, image_column, image_url)

                                        if is_valid_gid(image_gid):
                                            remember_file_reference(existing_files, filename, image_gid, image_url)
                                        else:
                                            print(
                                                f"⚠️ Unable to determine gid for image {filename}; variant metafields will "
                                                "re-upload if needed."
                                            )
                                    else:
                                        print(f"Failed to upload image {filename} to product {product_id}: {response.status_code}, {response.text}")
                                else:
                                    print(f"Product ID missing for row {index}, cannot upload image {filename}")
                            else:
                                print(f"Image file {filename} not found in local folder.")

            


        # Initialize a variable to store the last valid handle
        last_valid_handle = None
        # Cache to avoid uploading the same variant image multiple times per product/option1
        variant_image_cache = {}
        # Proceed with updating products and variants
        for index, row in df.iterrows():
            product_id = row.get('ID')
            variant_id = row.get('Variant ID')
            handle = row.get('Handle')  # Retrieve the handle from the spreadsheet
            sku = row.get('Variant SKU')  # Retrieve the handle from the spreadsheet

            print(f"Processing Product: Title='{sku}', Handle='{handle}'")


            if pd.isna(product_id):
                product_id = None
            else:
                # Convert to string without the .0
                product_id = str(int(product_id))

            if pd.isna(variant_id):
                variant_id = None
            else:
                # Convert to string without the .0
                variant_id = str(int(variant_id))

            handle = None if pd.isna(handle) else handle
  


            if pd.notna(handle):
                original_handle_value = str(handle)
                handle_candidate = original_handle_value.split(".")[0].lower().replace(" ", "-").replace("/", "-")
                handle_candidate = re.sub(r"[^a-z0-9-]", "", handle_candidate)

                if handle_candidate:
                    handle = handle_candidate
                    last_valid_handle = handle  # Update the last valid handle if the current row has one
                elif last_valid_handle:
                    print(
                        f"Detected invalid handle '{original_handle_value}'. Reusing last known valid handle '{last_valid_handle}'."
                    )
                    handle = last_valid_handle
                else:
                    print(f"Detected invalid handle '{original_handle_value}' and no prior handle to reuse. Skipping row {index}.")
                    continue
                    print(f"Handle='{handle}'")


            elif not handle and last_valid_handle:
                handle = last_valid_handle
                print(f"Reusing last known handle '{handle}' for row {index}.")

            # Use SKU if handle is not available
            elif not handle and pd.notna(sku):
                print(f"No handle found. Using SKU '{sku}' as identifier.") 
            elif not handle:
                print(f"Skipping row {index} due to missing handle and SKU.")
                continue

        
            product_options = []
            if pd.notna(handle):  # Only aggregate options for rows with a Handle
                seen_option_names = set()
                for i in range(1, 4):  # Assuming a maximum of 3 options (Option1, Option2, Option3)
                    option_name = normalize_option_value(row.get(f"Option{i} Name"))
                    if option_name:
                        normalized_option_name = option_name.casefold()
                        if normalized_option_name in seen_option_names:
                            continue
                        seen_option_names.add(normalized_option_name)
                        # Aggregate all unique values for this option across rows with the same Handle
                        option_values = []
                        for candidate_value in df.loc[df['Handle'] == handle, f"Option{i} Value"].tolist():
                            normalized_value = normalize_option_value(candidate_value)
                            if normalized_value and normalized_value not in option_values:
                                option_values.append(normalized_value)
                        if not option_values and i == 1:
                            option_values = ["Default Title"]
                        product_options.append({
                            "name": option_name,
                            "values": option_values
                        })

         
            variant_name_parts = []
            for i in range(1, 4):
                normalized_variant_option = normalize_option_value(row.get(f'Option{i} Value'))
                if normalized_variant_option:
                    variant_name_parts.append(normalized_variant_option)
            variant_name = " / ".join(variant_name_parts) or "Default Title"

            # Skip invalid or missing Product Title and Variant Name
            if not has_cell_value(row.get('Title')) and not variant_name:
                print(f"Skipping row {index} due to missing product title and variant name.")
                continue

            metafields = collect_metafields_from_row(row, 'Metafield:')
            variant_metafields = collect_metafields_from_row(row, 'Variant Metafield:')

            # Prepare the product update data if the Product ID is available
            if has_cell_value(row.get('Title')):
                variant_payload = {
                    "id": variant_id,
                    "price": row['Variant Price'],
                    "inventory_policy": "continue" if not has_cell_value(row.get("Variant Inventory Qty")) else "deny",
                    "inventory_quantity": int(row["Variant Inventory Qty"]) if has_cell_value(row.get("Variant Inventory Qty")) else None,
                    "inventory_management": "shopify" if has_cell_value(row.get("Variant Inventory Qty")) else None
                }

                for i in range(1, 4):
                    normalized_option_value = normalize_option_value(row.get(f'Option{i} Value'))
                    if normalized_option_value is not None:
                        variant_payload[f"option{i}"] = normalized_option_value

                if has_cell_value(row.get('Variant SKU')):
                    variant_payload["sku"] = row.get('Variant SKU')

                if has_cell_value(row.get("Variant Barcode")):
                    variant_payload["barcode"] = str(row.get("Variant Barcode")).split(".")[0]

                if has_cell_value(row.get('Variant Weight')):
                    variant_payload["weight"] = row.get('Variant Weight')

                if has_cell_value(row.get('Variant Weight Unit')):
                    variant_payload["weight_unit"] = row.get('Variant Weight Unit')

                product_data = {
                    "product": {
                        "id": product_id,
                        "title": row['Title'],
                        "options": product_options or [{"name": "Title", "values": ["Default Title"]}],
                        "variants": [variant_payload]
                    }
                }

                # Add optional fields only if they exist or have valid values
                if has_cell_value(row.get('Body HTML')):
                    product_data["product"]["body_html"] = row['Body HTML']


                if has_cell_value(row.get('Type')):
                    product_data["product"]["product_type"] = row['Type']

                if has_cell_value(row.get('Template Suffix')):
                    product_data["product"]["template_suffix"] = row['Template Suffix']

                # Conditionally add 'vendor' if present and valid
                if has_cell_value(row.get('Vendor')):
                    product_data["product"]["vendor"] = row['Vendor']

                # Conditionally add 'tags' if present and valid
                if has_cell_value(row.get('Tags')):
                    product_data["product"]["tags"] = row['Tags']

                if has_cell_value(row.get('Status')):
                    product_data["product"]["status"] = row['Status']

                

            
                print(f"Updating product for handle '{handle}':", json.dumps(product_data, indent=4))
                time.sleep(1)  # Delay for half a second before retrying

                                
                if product_id and variant_id:
                    print(f"Updating product with ID '{product_id}' and variant ID '{variant_id}'...")
                    product_id, updated_variant_id = update_product(product_id, product_data)
                    if updated_variant_id:
                        variant_id = updated_variant_id
                elif handle:
                    print(f"No valid product or variant ID found. Updating by handle '{handle}'...")
                    product_id, variant_id = update_product_by_handle(handle, product_data)
                elif sku:
                    print(f"No handle found. Updating by SKU '{sku}'...")
                    product_id, variant_id = update_product_by_sku(sku, product_data)

                if product_id:
                    print(f"Product processed successfully with ID '{product_id}' and Variant ID '{variant_id}'.")
                    set_dataframe_cell(df, index, 'ID', int(product_id))
                    if variant_id:
                        set_dataframe_cell(df, index, 'Variant ID', int(variant_id))


                    # Handle market-specific pricing dynamically
                    for column, market_name in pricing_columns.items():
                        print(f"Processing column '{column}' for market '{market_name}'...")

                        if market_name in market_names:
                            print(f"Market '{market_name}' found in Shopify markets.")
                            price_amount = row.get(column)
                            print(f"Price from column '{column}': {price_amount}")

                            price_list_id = get_price_list_id_for_market(market_name)
                            print(f"Price list ID for market '{market_name}': {price_list_id}")

                            if price_list_id and price_amount:
                                print(f"Adding fixed price for market '{market_name}', variant ID '{variant_id}', price '{price_amount}'")
                                add_fixed_price_for_market(price_list_id, variant_id, price_amount)
                                print(f"Fixed price added successfully for market '{market_name}'.")
                            else:
                                if not price_list_id:
                                    print(f"[WARNING] Price list ID is missing for market '{market_name}'. Unable to add fixed price.")
                                if not price_amount:
                                    print(f"[WARNING] Price amount is missing in column '{column}' for row {index}. Skipping this market.")
                        else:
                            print(f"[INFO] Market '{market_name}' from column '{column}' not found in Shopify markets. Skipping.")


                else:
                    print("Failed to create a new product.")

                if product_id:
                    print("Updating Images and Metafields now.")

                    # Collect image URLs and alt texts
                    images = []
                    alt_texts = []
                    for i in range(1, 21):
                        image_column = f"Image {i}"
                        alt_column = f"Image {i} Alt"
                        image_url = row.get(image_column)
                        alt_text = row.get(alt_column)

                        if image_url and isinstance(image_url, str):
                            if image_url.startswith('http'):
                                # If it's already a URL, append it directly
                                images.append(image_url)
                                alt_texts.append(alt_text)
                            else:
                                # Handle local file upload (if necessary)
                                filename = image_url
                                file_path_local = os.path.join(IMAGE_FOLDER, filename)
                                if os.path.exists(file_path_local):
                                    # Upload the image and get the URL
                                    with open(file_path_local, "rb") as image_file:
                                        encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
                                    image_data = {
                                        "image": {
                                            "attachment": encoded_string,
                                            "filename": filename,
                                            "alt": alt_text if pd.notna(alt_text) else filename
                                        }
                                    }
                                    url_api = f"{BASE_URL}/products/{product_id}/images.json"
                                    response = requests.post(url_api, headers=headers, json=image_data)
                                    if response.status_code in [200, 201]:
                                        image = response.json().get('image', {})
                                        image_url = image.get('src')
                                        print(f"Successfully uploaded image {filename} to product {product_id}")
                                        # Update cell value with image URL
                                        set_dataframe_cell(df, index, image_column, image_url)
                                        images.append(image_url)
                                        alt_texts.append(alt_text)
                                    else:
                                        print(f"Failed to upload image {filename} to product {product_id}: {response.status_code}, {response.text}")
                                else:
                                    print(f"Image file {filename} not found in local folder.")
                        else:
                            continue  # Skip if image URL is not valid

                    # Update product images with alt texts
                    if images:
                        update_product_images(product_id, images, alt_texts)

                        # Retrieve the updated images to get their IDs
                        product_images = get_product_images(product_id, force_refresh=True)
                        if not product_images:
                            print(f"No images retrieved for product {product_id} after update.")

                    # Update metafields for the product
                    if metafields:
                        if pd.notna(product_id):
                            update_metafields(product_id, metafields, existing_files, index, df)
                        else:
                            print(f"Product ID missing for row {index}, cannot update metafields.")
                else:
                    print("Skipping product images and product metafields because the product update did not succeed.")

            # Prepare the variant update data if there is a variant
            if variant_name:
                variant_payload = {
                    "id": variant_id,
                    "price": row['Variant Price'],
                }

                for i in range(1, 4):
                    normalized_variant_option = normalize_option_value(row.get(f'Option{i} Value'))
                    if normalized_variant_option is not None:
                        variant_payload[f"option{i}"] = normalized_variant_option

                variant_data = {"variant": variant_payload}

                variant_image_value = row.get('Variant Image')
                variant_image_alt = row.get('Variant Image Alt')
                option1_value = row.get('Option1 Value')
                cache_key = None
                cached_entry = None
                cached_image_id = None
                cached_media_id = None
                normalized_identifier = None


                if isinstance(option1_value, str) and option1_value.strip():
                    option1_identifier = option1_value.strip().lower()
                elif pd.notna(option1_value):
                    option1_identifier = str(option1_value).strip().lower()
                else:
                    option1_identifier = None

                if handle:
                    product_identifier = handle
                else:
                    product_identifier = product_id
                if option1_identifier and product_identifier:
                    cache_key = (str(product_identifier), option1_identifier)
                    cached_entry = variant_image_cache.get(cache_key)

                if isinstance(variant_image_value, str):
                    normalized_identifier = variant_image_value.strip()
                elif pd.notna(variant_image_value):
                    normalized_identifier = str(variant_image_value).strip()

                image_id = None
                media_id = None


                if cached_entry:
                    cached_image_id = cached_entry.get("image_id")
                    cached_url = cached_entry.get("url")
                    cached_identifier = cached_entry.get("identifier")
                    cached_media_id = cached_entry.get("media_id")

                    if not normalized_identifier and cached_identifier:
                        normalized_identifier = cached_identifier

                    if cached_identifier == normalized_identifier or not normalized_identifier:
                        if cached_url:
                            set_dataframe_cell(df, index, 'Variant Image', cached_url)
                        image_id = cached_image_id
                        media_id = cached_media_id

                if not image_id and not media_id and has_cell_value(variant_image_value):
                    image_id, media_id = ensure_variant_image(
                        product_id, variant_image_value, variant_image_alt, index, handle
                    )
                    if (image_id or media_id) and cache_key:
                        resolved_url = df.at[index, 'Variant Image']
                        variant_image_cache[cache_key] = {
                            "image_id": image_id,
                            "media_id": media_id,
                            "url": resolved_url if resolved_url else None,
                            "identifier": normalized_identifier,
                        }
                elif (image_id or media_id) and cache_key and not cached_entry:
                    resolved_url = df.at[index, 'Variant Image']
                    variant_image_cache[cache_key] = {
                        "image_id": image_id,
                        "media_id": media_id,
                        "url": resolved_url if resolved_url else None,
                        "identifier": normalized_identifier,
                    }

                if image_id:
                    variant_data["variant"]["image_id"] = image_id

                if has_cell_value(row.get('Variant SKU')):
                    variant_data["variant"]["sku"] = row['Variant SKU']

                if has_cell_value(row.get('Variant Barcode')):
                    variant_data["variant"]["barcode"] = (
                        str(int(row["Variant Barcode"]))
                        if pd.notna(row.get("Variant Barcode")) and str(row.get("Variant Barcode")).strip() != ""
                        else None
                    )

                if has_cell_value(row.get('Variant Weight')):
                    variant_data["variant"]["weight"] = row['Variant Weight']

                if has_cell_value(row.get('Variant Weight Unit')):
                    variant_data["variant"]["weight_unit"] = row['Variant Weight Unit']

                variant_data["variant"]["inventory_policy"] = (
                    "continue" if not has_cell_value(row.get("Variant Inventory Qty")) else "deny"
                )

                if has_cell_value(row.get("Variant Inventory Qty")):
                    inventory_qty = int(row["Variant Inventory Qty"])
                    variant_data["variant"]["inventory_quantity"] = inventory_qty
                    variant_data["variant"]["inventory_management"] = "shopify" if inventory_qty > 0 else None





            print(f"Updating or creating variant for handle '{handle}':")
            time.sleep(1)  # Delay for half a second before retrying
            variant_id = update_or_create_variant_by_handle(handle, variant_data)

            if variant_id:
                set_dataframe_cell(df, index, 'Variant ID', int(variant_id))
                print(f"Variant processed successfully with ID: {variant_id}")
                # Proceed with additional logic using `variant_id`, like setting market-specific prices

                                # If we resolved a GraphQL media GID for the variant’s image, append it to the variant
                if media_id:
                    pid = get_product_id_by_handle(handle)
                    if pid:
                        append_media_to_variant(pid, variant_id, media_id)
                    else:
                        print(f"⚠️ Could not resolve product_id for handle '{handle}' to append media.")


                # Handle market-specific pricing dynamically
                for column, market_name in pricing_columns.items():
                    print(f"Processing column '{column}' for market '{market_name}'...")

                    if market_name in market_names:
                        print(f"Market '{market_name}' found in Shopify markets.")
                        price_amount = row.get(column)
                        print(f"Price from column '{column}': {price_amount}")

                        price_list_id = get_price_list_id_for_market(market_name)
                        print(f"Price list ID for market '{market_name}': {price_list_id}")

                        if price_list_id and price_amount:
                            print(f"Adding fixed price for market '{market_name}', variant ID '{variant_id}', price '{price_amount}'")
                            add_fixed_price_for_market(price_list_id, variant_id, price_amount)
                            print(f"Fixed price added successfully for market '{market_name}'.")
                        else:
                            if not price_list_id:
                                print(f"[WARNING] Price list ID is missing for market '{market_name}'. Unable to add fixed price.")
                            if not price_amount:
                                print(f"[WARNING] Price amount is missing in column '{column}' for row {index}. Skipping this market.")
                    else:
                        print(f"[INFO] Market '{market_name}' from column '{column}' not found in Shopify markets. Skipping.")

                if variant_metafields:
                    update_variant_metafields(variant_id, variant_metafields, existing_files, index, df)

                else:
                    print(f"Failed to process variant for handle '{handle}'.")




        # After processing all rows, save the updated DataFrame back to Excel
        df.to_excel(file_path, index=False)
        print("Spreadsheet updated with new image URLs.")

    # Function to prompt the user to select a file
    def get_file_path():
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx")]
        )
        return file_path

    # Example usage
    file_path = get_file_path()
    if file_path:
        upload_changes_from_spreadsheet(file_path)
    else:
        print("No file selected.")# Function to run the downloader logic



def collection_run_downloader_logic():
    # Get the directory where the executable or script is located
    if getattr(sys, 'frozen', False):  # If running as an EXE
        script_dir = os.path.dirname(sys.executable)
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))

    # Build the full path to 'credentials.txt'
    credentials_path = os.path.join(script_dir, 'credentials.txt')

    shopify_context = load_shopify_context(credentials_path)
    SHOP_NAME = shopify_context["shop_name"]
    ACCESS_TOKEN = shopify_context["access_token"]
    log_shopify_access_scope_diagnostics(shopify_context)

    # Shopify API URL
    BASE_URL = shopify_context["base_url"]
    HEADERS = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": ACCESS_TOKEN
    }

    def collection_fetch_all():
        collections = []
        collection_types = ["smart_collections", "custom_collections"]
        
        for collection_type in collection_types:
            url = f"{BASE_URL}/{collection_type}.json?limit=250"
            while url:
                response = requests.get(url, headers=HEADERS)
                if response.status_code != 200:
                    print(f"Error fetching {collection_type}: {response.status_code}")
                    break
                
                data = response.json().get(collection_type, [])
                collections.extend(data)
                
                # Handle pagination
                link_header = response.headers.get("Link")
                next_url = None
                if link_header:
                    links = link_header.split(',')
                    for link in links:
                        if 'rel="next"' in link:
                            next_url = link.split(';')[0].strip('<> ')
                            break
                url = next_url
                time.sleep(1)
        
        return collections

    def collection_fetch_metafields(collection_id):
        metafields = []
        url = f"{BASE_URL}/collections/{collection_id}/metafields.json?limit=250"
        
        response = requests.get(url, headers=HEADERS)
        if response.status_code == 200:
            metafields = response.json().get("metafields", [])
        else:
            print(f"Error fetching metafields for collection {collection_id}: {response.status_code}")
        
        return metafields

    def collection_fetch_products(collection_id):
        products = []
        url = f"{BASE_URL}/collects.json?collection_id={collection_id}&limit=250"
        while url:
            response = requests.get(url, headers=HEADERS)
            if response.status_code != 200:
                print(f"Error fetching products for collection {collection_id}: {response.status_code}")
                break
            
            data = response.json().get("collects", [])
            products.extend([item["product_id"] for item in data])
            
            link_header = response.headers.get("Link")
            next_url = None
            if link_header:
                links = link_header.split(',')
                for link in links:
                    if 'rel="next"' in link:
                        next_url = link.split(';')[0].strip('<> ')
                        break
            url = next_url
            time.sleep(1)
        
        return products

    def collection_save_to_excel(collections, metafields_data, collection_products):
        print("Processing data and preparing to save to Excel...")
        data = []
        all_metafield_keys = set()
        
        for collection in collections:
            print(f"Processing collection: {collection['title']}")
            collection_type = "Smart" if "rules" in collection else "Manual"
            metafields = [mf for mf in metafields_data if mf[0] == collection["id"]]
            products = [cp[1] for cp in collection_products if cp[0] == collection["id"]]
            
            rules = collection.get("rules", [])
            rule_conditions = ", ".join([f"{rule['column']} {rule['relation']} {rule['condition']}" for rule in rules])
            
            collection_data = {
                "ID": collection["id"],
                "Title": collection["title"],
                "Handle": collection["handle"],
                "Collection Type": collection_type,
                "Created At": collection.get("created_at", ""),
                "Updated At": collection.get("updated_at", ""),
                "Image Src": collection.get("image", {}).get("src", ""),
                "Products": ", ".join(map(str, products)),
                "Conditions": rule_conditions
            }
            
            for mf in metafields:
                column_name = f"Metafield: {mf[1]}.{mf[2]} [{mf[4]}]"
                all_metafield_keys.add(column_name)
                collection_data[column_name] = mf[3]
            
            data.append(collection_data)
        
        df = pd.DataFrame(data)
        for key in all_metafield_keys:
            if key not in df.columns:
                df[key] = None
                
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = f"shopify_collections_{current_time}.xlsx"

        df.to_excel(file_path, index=False)
        
        wb = load_workbook(file_path)
        ws = wb.active
        bold_font = Font(bold=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].value:
                for cell in row:
                    cell.font = bold_font
        ws.freeze_panes = ws['B2']
        wb.save(file_path)
        
        print("Data has been saved successfully.")

    print("Fetching collections...")
    collections = collection_fetch_all()
    
    print("Fetching metafields and products for each collection...")
    metafields_data = []
    collection_products = []
    
    for col in collections:
        col_id = col["id"]
        metafields = collection_fetch_metafields(col_id)
        for mf in metafields:
            metafields_data.append([col_id, mf["namespace"], mf["key"], mf["value"], mf.get("type", "")])
        
        products = collection_fetch_products(col_id)
        for prod in products:
            collection_products.append([col_id, prod])
    
    print("Saving data to Excel...")
    collection_save_to_excel(collections, metafields_data, collection_products)
    print("Collection download completed.")

def collection_run_uploader_logic():
    # Get script directory
    if getattr(sys, 'frozen', False):
        script_dir = os.path.dirname(sys.executable)
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))

    # Load credentials
    credentials_path = os.path.join(script_dir, 'credentials.txt')
    shopify_context = load_shopify_context(credentials_path)
    SHOP_NAME = shopify_context["shop_name"]
    ACCESS_TOKEN = shopify_context["access_token"]
    log_shopify_access_scope_diagnostics(shopify_context)

    # Shopify API URLs
    BASE_URL = shopify_context["base_url"]
    GRAPHQL_URL = shopify_context["graphql_url"]

    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": ACCESS_TOKEN
    }

    graphql_headers = headers.copy()

    IMAGE_FOLDER = os.path.join(script_dir, 'img')  # Adjust as needed
    FILE_FOLDER = os.path.join(script_dir, 'files')

    asset_directories = []
    for candidate in [IMAGE_FOLDER, FILE_FOLDER, script_dir]:
        if candidate and candidate not in asset_directories:
            asset_directories.append(candidate)

    def resolve_local_asset_path(filename):
        return resolve_asset_from_directories(filename, asset_directories)
    
    def create_collection(collection_data, is_smart=False):
        url = f"{BASE_URL}/smart_collections.json" if is_smart else f"{BASE_URL}/custom_collections.json"
        print(f"📡 Creating {'Smart' if is_smart else 'Manual'} Collection: {collection_data}")
        response = requests.post(url, headers=headers, json=collection_data)
        
        print(f"📩 Response: {response.status_code} - {response.text}")

        if response.status_code in [200, 201]:
            return response.json().get("smart_collection" if is_smart else "custom_collection", {})
        else:
            print(f"❌ Failed to create collection: {response.status_code}, {response.text}")
            return None

    def delete_metafield(collection_id, metafield_id):
        """Deletes a metafield from a collection."""
        url = f"{BASE_URL}/metafields/{metafield_id}.json"
        print(f"🗑 Deleting metafield {metafield_id} from collection {collection_id}...")
        response = requests.delete(url, headers=headers)
        
        print(f"📩 Delete Response: {response.status_code}")

        if response.status_code == 200:
            print(f"✅ Deleted metafield {metafield_id} for collection {collection_id}")
        else:
            print(f"❌ Failed to delete metafield {metafield_id}: {response.status_code}, {response.text}")

    # Function to create a staged upload
    def staged_upload_create(filename, mime_type):
        query = """
        mutation {
        stagedUploadsCreate(input: {
            resource: FILE,
            filename: "%s",
            mimeType: "%s",
            httpMethod: POST
        }) {
            stagedTargets {
            url
            parameters {
                name
                value
            }
            }
            userErrors {
            field
            message
            }
        }
        }
        """ % (filename, mime_type)

        response = requests.post(GRAPHQL_URL, json={"query": query}, headers=graphql_headers)
        data = response.json()
        if "data" in data and "stagedUploadsCreate" in data["data"]:
            return data["data"]["stagedUploadsCreate"]["stagedTargets"][0]
        else:
            print("Error in staged upload creation:", data)
            return None

    # Function to upload the file to the staging URL
    def upload_file_to_staging(staging_target, file_path):
        url = staging_target["url"]
        files = {"file": open(file_path, "rb")}
        form_data = {param["name"]: param["value"] for param in staging_target["parameters"]}
        response = requests.post(url, data=form_data, files=files)

        # Parsing the XML response to get the file location
        if response.status_code == 201:
            print(f"File {file_path} successfully uploaded to staging URL.")
            xml_response = ET.fromstring(response.text)
            location_url = xml_response.find('Location').text  # Extracting Location URL from XML
            return location_url
        else:
            print(f"Failed to upload file. Status: {response.status_code}, {response.text}")
            return None

    # Function to commit the file with fileCreate
    def commit_file_to_shopify(file_name, original_source):
        query = """
        mutation fileCreate($files: [FileCreateInput!]!) {
        fileCreate(files: $files) {
            files {
            id  # Fetch the gid after file creation
            alt
            createdAt
            ... on GenericFile {
                url
            }
            ... on MediaImage {
                image {
                url
                }
            }
            }
            userErrors {
            code
            field
            message
            }
        }
        }
        """
        variables = {
            "files": [
                {
                    "alt": file_name,
                    "originalSource": original_source
                }
            ]
        }
        response = requests.post(GRAPHQL_URL, json={"query": query, "variables": variables}, headers=graphql_headers)
        data = response.json()
        if "data" in data and "fileCreate" in data["data"]:
            file_info = data["data"]["fileCreate"]["files"]
            gid = file_info[0]["id"]
            print(f"File {file_name} successfully committed to Shopify.")
            print(f"GID: {gid}")  # Print the gid here
            return file_info
        else:
            print("Error in file commit:", data)
            return None

    # Function to get image URL for a given GID
    def get_image_url_for_gid(gid):
        query = f"""
        {{
        node(id: "{gid}") {{
            ... on GenericFile {{
            url
            }}
            ... on MediaImage {{
            image {{
                url
            }}
            }}
        }}
        }}
        """
        response = requests.post(GRAPHQL_URL, json={"query": query}, headers=graphql_headers)
        if response.status_code == 200:
            data = response.json()
            node = (((data or {}).get('data') or {}).get('node') or {})
            url = None
            if 'url' in node:
                url = node['url']
            elif 'image' in node and node['image']:
                url = node['image'].get('url')
            return url
        else:
            print(f"Failed to get image URL for GID {gid}")
            return None

    def upload_image_to_shopify(file_path):
        filename = os.path.basename(file_path)

        # Normalize the filename before accessing it
        normalized_filename = normalize_filename(filename)
        folder_path = os.path.dirname(file_path)

        file_path = os.path.join(os.path.dirname(file_path), normalized_filename)

        
        # Check if file exists after normalization
        if not os.path.exists(file_path):
            print(f"Image file {normalized_filename} not found in local folder.")
            return None, None


          # Ensure the filename is URL-encoded
        encoded_filename = encode_filename(filename)    
    
    
        # Extract the file name and MIME type
        mime_type = guess_mime_type(filename)

        # Step 1: Create the staged upload
        staging_target = staged_upload_create(filename, mime_type)
        if not staging_target:
            return None, None

        # Step 2: Upload the file to the staging URL
        location_url = upload_file_to_staging(staging_target, file_path)
        if not location_url:
            return None, None

        # Step 3: Commit the file to Shopify
        file_info = commit_file_to_shopify(filename, location_url)
        if file_info:
            # file_info contains the files array
            file_data = file_info[0]
            gid = file_data['id']
            # Get the URL
            url = None
            if 'url' in file_data:
                url = file_data['url']
            elif 'image' in file_data and file_data['image']:
                url = file_data['image'].get('url')
            if not url:
                # Get URL via get_image_url_for_gid
                url = get_image_url_for_gid(gid)
                if not url:
                    print(f"⚠️ Unable to resolve URL for uploaded file with GID {gid}")

            print(f"✅ Image uploaded successfully: GID={gid}, URL={url}")

            return url, gid 
        else:
            return None, None

    def get_all_files():
        """Retrieves all existing file references in Shopify."""
        print("📡 Fetching all existing Shopify files...")
        all_files = {}
        has_next_page = True
        cursor = None

        while has_next_page:
            query = f"""
            {{
            files(first: 250{' , after: "' + cursor + '"' if cursor else ''}) {{
                edges {{
                node {{
                    id
                    alt
                    ... on GenericFile {{
                    url
                    }}
                    ... on MediaImage {{
                    image {{
                        url
                    }}
                    }}
                }}
                cursor
                }}
                pageInfo {{
                hasNextPage
                }}
            }}
            }}
            """
            response = requests.post(GRAPHQL_URL, json={"query": query}, headers=headers)
            if response.status_code == 200:
                data = response.json()
                for file in ((((data or {}).get("data") or {}).get("files")) or {}).get("edges", []):
                    node = file["node"]
                    gid = node["id"]
                    alt = node["alt"]
                    url = node.get('url') or node.get('image', {}).get('url')
                    if alt:
                        remember_file_reference(all_files, alt, gid, url)
                    cursor = file["cursor"]

                has_next_page = data["data"]["files"]["pageInfo"]["hasNextPage"]
                print(f"✅ {len(all_files)} files retrieved so far...")
            else:
                print(f"❌ Error fetching files. Response: {response.status_code}")
                return None

        print(f"✅ Finished fetching {len(all_files)} existing files.")
        return all_files

    def update_metafields(collection_id, metafields, existing_files, row_index, df):
        """Handles metafield updates for collections, including file uploads and deletions."""
        if not collection_id:
            print(f"⚠️ Skipping metafield update: Missing collection ID.")
            return

        print(f"\n📡 Fetching existing metafields for Collection ID: {collection_id}...")
        url = f"{BASE_URL}/collections/{collection_id}/metafields.json"
        response = requests.get(url, headers=headers)
        current_metafields = response.json().get('metafields', []) if response.status_code == 200 else []
        current_metafields_dict = {f"{mf['namespace']}.{mf['key']}": mf['id'] for mf in current_metafields}

        print(f"🔍 Found {len(current_metafields)} existing metafields.")

        for column, value in metafields.items():
            key_type_str = column.replace('Metafield: ', '').split(' ')
            key = key_type_str[0]
            field_type = key_type_str[1].replace('[', '').replace(']', '') if len(key_type_str) > 1 else 'single_line_text_field'

            namespace, key = key.split('.')

            print(f"\n📝 Processing metafield → Namespace: '{namespace}', Key: '{key}', Type: '{field_type}', Value: {value}")

            if pd.isna(value) or value is None:
                metafield_key = f"{namespace}.{key}"
                if metafield_key in current_metafields_dict:
                    print(f"🗑️ Deleting metafield '{namespace}.{key}' (ID: {current_metafields_dict[metafield_key]}) from collection {collection_id}...")
                    delete_metafield(collection_id, current_metafields_dict[metafield_key])
                else:
                    print(f"⚠️ Metafield '{namespace}.{key}' does not exist, skipping deletion.")
                continue

            metafield_data = None

            if field_type == 'file_reference':
                value_gid = None
                filename = None
                if isinstance(value, str):
                    candidate = value.strip()
                    if candidate.startswith('gid://'):
                        value_gid = candidate
                        print(f"✅ Using existing GID: {value_gid}")
                    elif candidate.startswith('http'):
                        filename = extract_filename_from_value(candidate)
                        existing_entry = fetch_file_reference(existing_files, filename)
                        value_gid = existing_entry[0] if existing_entry else None
                        print(f"🔍 Retrieved GID from existing files: {value_gid}")
                    else:
                        filename = extract_filename_from_value(candidate)
                        existing_entry = fetch_file_reference(existing_files, filename)
                        if existing_entry:
                            value_gid = existing_entry[0]
                            set_dataframe_cell(df, row_index, column, existing_entry[1] or value_gid)
                            print(f"✅ Found existing upload: {candidate}, GID: {value_gid}")
                        else:
                            file_path_local = resolve_local_asset_path(candidate)
                            if file_path_local:
                                print(f"📤 Uploading {candidate} to Shopify...")
                                uploaded_url, gid = upload_image_to_shopify(file_path_local)
                                if gid and uploaded_url:
                                    remember_file_reference(existing_files, filename or candidate, gid, uploaded_url)
                                    value_gid = gid
                                    set_dataframe_cell(df, row_index, column, uploaded_url)
                                    print(f"✅ Successfully uploaded '{candidate}', New GID: {value_gid}")
                                else:
                                    print(f"❌ Failed to upload file '{candidate}', skipping metafield update.")
                                    continue
                            else:
                                print(f"⚠️ File '{candidate}' not found in local folders, skipping metafield update.")
                                continue
                else:
                    print(f"⚠️ Skipping non-string value for metafield {namespace}.{key}.")
                    continue

                if value_gid:
                    metafield_data = {
                        "metafield": {
                            "namespace": namespace,
                            "key": key,
                            "value": value_gid,
                            "type": 'file_reference'
                        }
                    }
                else:
                    print(f"❌ Cannot determine GID for metafield value '{value}'.")
                    continue

            elif field_type == 'list.file_reference':
                if isinstance(value, str):
                    file_names = [filename.strip() for filename in value.split(',') if filename.strip()]
                    file_gids = []

                    print(f"🔍 Processing metafield '{key}'  with {len(file_names)} files: {file_names}")

                    for raw_name in file_names:
                        value_gid = None
                        filename = extract_filename_from_value(raw_name)

                        if raw_name.startswith('gid://'):
                            value_gid = raw_name
                            print(f"✅ Using existing GID: {value_gid}")
                        elif raw_name.startswith('http'):
                            existing_entry = fetch_file_reference(existing_files, filename)
                            value_gid = existing_entry[0] if existing_entry else None
                            print(f"🔍 Retrieved GID from existing files for {filename}: {value_gid}")
                        else:
                            existing_entry = fetch_file_reference(existing_files, filename)
                            if existing_entry:
                                value_gid = existing_entry[0]
                                set_dataframe_cell(df, row_index, column, existing_entry[1] or value_gid)
                                print(f"✅ Found existing upload for {raw_name}, GID: {value_gid}")
                            else:
                                file_path_local = resolve_local_asset_path(raw_name)
                                if file_path_local:
                                    print(f"📤 Uploading {raw_name} to Shopify...")
                                    uploaded_url, gid = upload_image_to_shopify(file_path_local)
                                    if gid and uploaded_url:
                                        remember_file_reference(existing_files, filename or raw_name, gid, uploaded_url)
                                        value_gid = gid
                                        set_dataframe_cell(df, row_index, column, uploaded_url)
                                        print(f"✅ Successfully uploaded {raw_name}, new GID: {value_gid}")
                                    else:
                                        print(f"❌ Failed to upload file {raw_name}")
                                        continue
                                else:
                                    print(f"⚠️ File {raw_name} not found locally.")
                                    continue

                        if value_gid:
                            file_gids.append(value_gid)
                        else:
                            print(f"❌ Cannot find GID for file {raw_name}")

                    print(f"📝 Final GID list for metafield '{key}': {file_gids}")

                    if file_gids:
                        metafield_data = {
                            "metafield": {
                                "namespace": namespace,
                                "key": key,
                                "value": json.dumps(file_gids),
                                "type": 'list.file_reference'
                            }
                        }
                        print(f"📡 Sending metafield update to Shopify: {metafield_data}")
                    else:
                        print(f"❌ Skipping metafield update for '{key}' because the file list is empty (avoiding 422 error).")
                        continue
                else:
                    print(f"⚠️ Skipping non-string value for metafield {key}.")
                    continue

            elif field_type == 'url':
                if isinstance(value, str) and value.strip():
                    raw_value = value.strip()
                    final_url = None
                    filename = extract_filename_from_value(raw_value)

                    if raw_value.startswith('http'):
                        existing_entry = fetch_file_reference(existing_files, filename)
                        final_url = existing_entry[1] if existing_entry and existing_entry[1] else raw_value
                        print(f"🔍 Using provided URL for metafield '{key}': {final_url}")
                    else:
                        existing_entry = fetch_file_reference(existing_files, filename)
                        if existing_entry and existing_entry[1]:
                            final_url = existing_entry[1]
                            print(f"✅ Reusing existing upload URL for '{filename}'")
                        else:
                            file_path_local = resolve_local_asset_path(raw_value)
                            if not file_path_local:
                                print(f"⚠️ File {raw_value} not found for URL metafield {namespace}.{key}")
                                continue
                            uploaded_url, gid = upload_image_to_shopify(file_path_local)
                            if uploaded_url:
                                remember_file_reference(existing_files, filename or raw_value, gid, uploaded_url)
                                final_url = uploaded_url
                                print(f"✅ Uploaded '{raw_value}' for URL metafield {namespace}.{key}")
                            else:
                                print(f"❌ Failed to upload file {raw_value} for URL metafield {namespace}.{key}")
                                continue

                    if not final_url:
                        print(f"⚠️ Skipping URL metafield {namespace}.{key} due to missing URL value.")
                        continue

                    set_dataframe_cell(df, row_index, column, final_url)

                    metafield_data = {
                        "metafield": {
                            "namespace": namespace,
                            "key": key,
                            "value": final_url,
                            "type": 'url'
                        }
                    }
                else:
                    print(f"⚠️ Skipping invalid URL metafield value for {namespace}.{key}")
                    continue

            else:
                formatted_value = format_metafield_text_value(value)
                print(f"📌 Adding text metafield: '{formatted_value}'")
                metafield_data = {
                    "metafield": {
                        "namespace": namespace,
                        "key": key,
                        "value": formatted_value,
                        "type": field_type.strip()
                    }
                }

            if metafield_data:
                url = f"{BASE_URL}/collections/{collection_id}/metafields.json"
                print(f"📡 Sending metafield update to Shopify API: {url}")
                print(f"📝 Payload: {json.dumps(metafield_data, indent=2, ensure_ascii=False)}")
                response = requests.post(url, headers=headers, json=metafield_data)
                if response.status_code in [200, 201]:
                    print(f"✅ Successfully updated metafield '{namespace}.{key}' for collection {collection_id}")
                else:
                    print(f"❌ Failed to update metafield '{namespace}.{key}' for collection {collection_id}: {response.status_code}")
                    print(f"⚠️ Response Body: {response.text}")
    def upload_collections_from_file(file_path):
        print(f"📂 Reading collections from file: {file_path}")
        df = pd.read_excel(file_path)
        df = df.where(pd.notnull(df), None)
        existing_files = get_all_files() or {}

        for _, row in df.iterrows():
            title = row["Title"]
            handle = row["Handle"]
            published = str(row.get("Published", "yes")).strip().lower() == "yes"

            # ✅ Extract conditions for smart collections
            conditions = []
            if "Conditions" in df.columns and pd.notna(row["Conditions"]):
                condition_strings = str(row["Conditions"]).split(";")
                for condition in condition_strings:
                    parts = condition.strip().split()
                    if len(parts) >= 3:
                        column, relation, condition_value = parts[0], parts[1], " ".join(parts[2:])
                        conditions.append({
                            "column": column,
                            "relation": relation,
                            "condition": condition_value
                        })

            is_smart = bool(conditions)

            print(f"🔍 Checking for existing collection with handle: {handle}")
            existing = find_existing_collection_by_handle(handle)

            if existing:
                collection_id = existing['id']
                print(f"🔄 Updating existing collection with ID {collection_id}")

                if is_smart:
                    collection_data = {
                        "smart_collection": {
                            "id": collection_id,
                            "title": title,
                            "handle": handle,
                            "published": published,
                            "rules": conditions,
                            "disjunctive": False
                        }
                    }
                else:
                    collection_data = {
                        "custom_collection": {
                            "id": collection_id,
                            "title": title,
                            "handle": handle,
                            "published": published
                        }
                    }

                success = update_collection(collection_id, collection_data, is_smart)
                if success:
                    print(f"✅ Updated Collection: {title} (ID: {collection_id})")
                    metafields = {col: row[col] for col in df.columns if col.startswith("Metafield:")}
                    if metafields:
                        update_metafields(collection_id, metafields, existing_files, _, df)
                else:
                    print(f"❌ Failed to update collection: {title}")
                continue  # Skip creation

            # ✅ If no existing collection, proceed to create a new one
            print(f"🆕 Creating {'Smart' if is_smart else 'Manual'} Collection: Title={title}, Handle={handle}, Published={published}")
            print(f"🛠 Conditions: {conditions if is_smart else 'None (Manual Collection)'}")

            if is_smart:
                collection_data = {
                    "smart_collection": {
                        "title": title,
                        "handle": handle,
                        "published": published,
                        "rules": conditions,
                        "disjunctive": False
                    }
                }
            else:
                collection_data = {
                    "custom_collection": {
                        "title": title,
                        "handle": handle,
                        "published": published
                    }
                }

            collection = create_collection(collection_data, is_smart)
            if collection:
                collection_id = collection.get("id")
                print(f"✅ Created Collection: {title} (ID: {collection_id})")
                metafields = {col: row[col] for col in df.columns if col.startswith("Metafield:")}
                if metafields:
                    update_metafields(collection_id, metafields, existing_files, _, df)
            else:
                print(f"❌ Failed to create collection: {title}")

    print("🎉 Collections upload completed.")

    def update_collection(collection_id, collection_data, is_smart=False):
        url = f"{BASE_URL}/smart_collections/{collection_id}.json" if is_smart else f"{BASE_URL}/custom_collections/{collection_id}.json"
        print(f"🛠 Updating {'Smart' if is_smart else 'Manual'} Collection ID {collection_id}")
        response = requests.put(url, headers=headers, json=collection_data)
        print(f"📩 Update Response: {response.status_code} - {response.text}")
        return response.status_code in [200, 201]

    def find_existing_collection_by_handle(handle):
        """Returns collection object (manual or smart) by handle, or None."""
        url = f"{BASE_URL}/custom_collections.json?handle={handle}"
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            collections = response.json().get('custom_collections', [])
            if collections:
                return collections[0]
        
        # If not found in manual, try smart
        url = f"{BASE_URL}/smart_collections.json?handle={handle}"
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            collections = response.json().get('smart_collections', [])
            if collections:
                return collections[0]
        
        return None


    def get_file_path():
        root = tk.Tk()
        root.withdraw()
        return filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx")])

    file_path = get_file_path()
    if file_path:
        upload_collections_from_file(file_path)
    else:
        print("❌ No file selected.")

def start_download():
    def after_download():
        download_button.config(state=tk.NORMAL)
        upload_button.config(state=tk.NORMAL)

    download_button.config(state=tk.DISABLED)
    upload_button.config(state=tk.DISABLED)

    # Start download in a separate thread
    thread = build_safe_thread("Download", run_downloader_logic)
    thread.start()

    # Wait for the thread to finish and then re-enable buttons
    root.after(100, check_thread, thread, after_download, "Download completed!")

def start_upload():
    def after_upload():
        download_button.config(state=tk.NORMAL)
        upload_button.config(state=tk.NORMAL)

    download_button.config(state=tk.DISABLED)
    upload_button.config(state=tk.DISABLED)

    # Start upload in a separate thread
    thread = build_safe_thread("Upload", run_uploader_logic)
    thread.start()

    # Wait for the thread to finish and then re-enable buttons
    root.after(100, check_thread, thread, after_upload, "Upload completed!")

def start_collection_download():
    def after_collection_download():
        collection_download_button.config(state=tk.NORMAL)
        collection_upload_button.config(state=tk.NORMAL)
    
    collection_download_button.config(state=tk.DISABLED)
    collection_upload_button.config(state=tk.DISABLED)
    
    thread = build_safe_thread("Collection download", collection_run_downloader_logic)
    thread.start()
    root.after(100, check_thread, thread, after_collection_download, "Download completed!")

def start_collection_upload():
    def after_collection_upload():
        collection_download_button.config(state=tk.NORMAL)
        collection_upload_button.config(state=tk.NORMAL)
    
    collection_download_button.config(state=tk.DISABLED)
    collection_upload_button.config(state=tk.DISABLED)
    
    thread = build_safe_thread("Collection upload", collection_run_uploader_logic)
    thread.start()
    root.after(100, check_thread, thread, after_collection_upload, "Upload completed!")

def download_shopify_files_alt_texts():
    import requests
    import pandas as pd
    import os
    from datetime import datetime

    if getattr(sys, 'frozen', False):
        script_dir = os.path.dirname(sys.executable)
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))
    
    shopify_context = load_shopify_context(os.path.join(script_dir, 'credentials.txt'))
    SHOP_NAME = shopify_context["shop_name"]
    ACCESS_TOKEN = shopify_context["access_token"]

    GRAPHQL_URL = shopify_context["graphql_url"]
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": ACCESS_TOKEN
    }

    print("📥 Fetching all uploaded Shopify files...")
    all_files = []
    has_next_page = True
    cursor = None

    while has_next_page:
        query = f"""
        {{
          files(first: 250 {', after: "' + cursor + '" ' if cursor else ''}) {{
            edges {{
              node {{
                id
                alt
                __typename
                ... on GenericFile {{
                  url
                }}
                ... on MediaImage {{
                  image {{
                    url
                  }}
                }}
              }}
              cursor
            }}
            pageInfo {{
              hasNextPage
            }}
          }}
        }}
        """

        response = requests.post(GRAPHQL_URL, headers=headers, json={"query": query})
        data = response.json()

        if "errors" in data:
            print(f"❌ Error fetching files: {data['errors']}")
            break

        files = data['data']['files']['edges']
        for file in files:
            node = file['node']
            gid = node['id']
            alt = node.get('alt', '')
            url = node.get('url') or (node.get('image', {}).get('url') if node.get('image') else None)
            filename = os.path.basename(url) if url else ""

            all_files.append({
                "Filename": filename,
                "Alt Text": alt,
                "GID": gid,
                "URL": url
            })

        has_next_page = data['data']['files']['pageInfo']['hasNextPage']
        if has_next_page:
            cursor = files[-1]['cursor']
    
    # Save to Excel
    df = pd.DataFrame(all_files)
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"shopify_uploaded_files_alt_texts_{current_time}.xlsx"
    df.to_excel(file_name, index=False)

    print(f"✅ Files Alt Texts exported to {file_name}")

def upload_shopify_files_alt_texts():
    import requests
    import pandas as pd
    import os
    import tkinter as tk
    from tkinter import filedialog

    if getattr(sys, 'frozen', False):
        script_dir = os.path.dirname(sys.executable)
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))

    shopify_context = load_shopify_context(os.path.join(script_dir, 'credentials.txt'))
    SHOP_NAME = shopify_context["shop_name"]
    ACCESS_TOKEN = shopify_context["access_token"]

    GRAPHQL_URL = shopify_context["graphql_url"]
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": ACCESS_TOKEN
    }

    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Select Shopify Files Alt Text Excel File",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not file_path:
        print("❌ No file selected.")
        return

    df = pd.read_excel(file_path)
        
    # 🛠 FIX: Replace NaN with empty string to avoid JSON errors
    df = df.fillna('')

    if "GID" not in df.columns or "New Alt Text" not in df.columns:
        print("❌ Excel must have 'GID' and 'New Alt Text' columns.")
        return

    for idx, row in df.iterrows():
        gid = row['GID']
        new_alt_text = row['New Alt Text']

        if not gid or not new_alt_text:
            print(f"⚠️ Skipping row {idx} due to missing GID or New Alt Text.")
            continue

        mutation = """
        mutation fileUpdate($id: ID!, $alt: String!) {
          fileUpdate(files: {id: $id, alt: $alt}) {
            files {
              id
              alt
            }
            userErrors {
              field
              message
            }
          }
        }
        """

        variables = {
            "id": gid,
            "alt": new_alt_text
        }

        response = requests.post(GRAPHQL_URL, headers=headers, json={"query": mutation, "variables": variables})
        result = response.json()

        if "errors" in result:
            print(f"❌ Error updating file {gid}: {result['errors']}")
        elif result['data']['fileUpdate']['userErrors']:
            print(f"❌ Shopify User Errors for file {gid}: {result['data']['fileUpdate']['userErrors']}")
        else:
            print(f"✅ Successfully updated Alt Text for File {gid}")

    print("✅ All Alt Text updates completed.")

def generate_seo_alt_texts():
    if getattr(sys, 'frozen', False):
        script_dir = os.path.dirname(sys.executable)
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))

    credentials = read_credentials(os.path.join(script_dir, 'credentials.txt'))
    openai_api_key = credentials['openai_api_key']
    store_name = credentials['store_name']

    client = OpenAI(api_key=openai_api_key)

    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Select Excel File with Image URLs",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not file_path:
        print("❌ No file selected.")
        return

    df = pd.read_excel(file_path)
    if "URL" not in df.columns or "Filename" not in df.columns or "Alt Text" not in df.columns:
        print("❌ Excel must have 'Filename', 'Alt Text', and 'URL' columns.")
        return

    df["New Alt Text"] = ""

    print(f"🔍 Processing {len(df)} images...")

    store_name = beautify_store_name(credentials['store_name'])

    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    final_output_file = f"shopify_images_with_new_alt_texts_{current_time}.xlsx"
    temp_file = f"temp_alt_texts_{current_time}.xlsx"
    backup_interval = 5

    for index, row in df.iterrows():
        image_url = row['URL']
        filename = row['Filename']
        old_alt_text = row['Alt Text']

        if not image_url or not isinstance(image_url, str):
            print(f"⚠️ Skipping row {index} due to missing image URL.")
            continue

        try:
            response = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "Du bist ein professioneller SEO-Experte für Alt-Texte. "
                            "Erstelle präzise, natürlich klingende Alt-Texte auf Deutsch, ideal für Suchmaschinenoptimierung (SEO). "
                            "Beschreibe den Bildinhalt klar und objektiv, mit Fokus auf sichtbare Objekte, Personen, Umgebung und Aktivitäten. "
                            "Begrenze den Alt-Text auf maximal 15 Wörter. "
                            "Verwende relevante Keywords und achte auf natürliche Sprache."
                        )
                    },
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": (
                                    f"Bildanalyse:\n"
                                    f"- Dateiname: {filename}\n"
                                    f"- Vorhandener Alt-Text: {old_alt_text}\n"
                                    f"- Markenname: {store_name}\n\n"
                                    f"Aufgabe:\n"
                                    "- Beschreibe präzise, was auf dem Bild zu sehen ist (z. B. Objekt(e), Personen, Tätigkeit, Umgebung), inklusive relevanter Merkmale wie Farbe, Material oder Funktion.\n"
                                    f"- Baue den Markennamen ({store_name}) sinnvoll und flüssig in den Text ein, falls passend.\n"
                                    "- Verwende thematisch relevante SEO-Begriffe (z. B. Lederstuhl, Turnringe, Waldtraining), wenn möglich natürlich integriert.\n"
                                    "- Maximal 15 Wörter, vollständiger deutscher Satz ohne Listenstil."
                                )
                            },
                            {
                                "type": "image_url",
                                "image_url": {"url": image_url}
                            }
                        ]
                    }
                ],
                temperature=0.2,
                max_tokens=100
            )

            generated_text = response.choices[0].message.content
            set_dataframe_cell(df, index, "New Alt Text", generated_text.strip())
            print(f"✅ Generated new alt text for row {index} – {generated_text.strip()}")

            # 🔁 Save backup every 5 rows
            if index % backup_interval == 0:
                try:
                    with file_lock:
                        df.to_excel(temp_file, index=False)
                        print(f"💾 Temp backup written at row {index}")
                except Exception as e:
                    print(f"⚠️ Backup write failed at row {index}: {e}")

        except Exception as e:
            print(f"❌ Error generating alt text for row {index}: {str(e)}")
            continue

    # ✅ Final save with retry logic
    for attempt in range(3):
        try:
            with file_lock:
                df.to_excel(final_output_file, index=False)
            print(f"✅ All done! File saved as {final_output_file}")
            break
        except PermissionError:
            print(f"❌ Attempt {attempt + 1}: Permission denied. Retrying in 2s...")
            time.sleep(2)
        except Exception as e:
            print(f"❌ Unexpected error during final save: {e}")
            break
    else:
        print(f"❌ Final save failed after 3 attempts. Temp file saved at: {temp_file}")

def beautify_store_name(store_name):
    try:
        store_name = get_shop_name(store_name)
    except RuntimeError:
        store_name = (store_name or "").strip()
    # Replace hyphens and underscores with spaces
    name = store_name.replace('-', ' ').replace('_', ' ')
    # Remove extra spaces (if any)
    name = ' '.join(name.split())
    # Capitalize first letter of each word
    name = name.title()
    return name

def check_thread(thread, callback, success_message=None):
    if thread.is_alive():
        # If the thread is still running, check again after 100 ms
        root.after(100, check_thread, thread, callback, success_message)
    else:
        callback()

        if getattr(thread, "task_failed", False):
            messagebox.showerror(
                "Error",
                getattr(thread, "error_message", "The operation failed."),
            )
            return

        if success_message:
            messagebox.showinfo("Success", success_message)


def build_safe_thread(task_name, target):
    def runner():
        current_thread = threading.current_thread()
        current_thread.task_failed = False
        current_thread.error_message = ""
        try:
            target()
        except Exception as exc:
            current_thread.task_failed = True
            current_thread.error_message = f"{task_name} failed: {exc}"
            traceback.print_exc()
            print(f"❌ {current_thread.error_message}")

    thread = threading.Thread(target=runner)
    thread.task_failed = False
    thread.error_message = ""
    return thread


def start_detached_task(task_name, target):
    build_safe_thread(task_name, target).start()

# Function to load the rest of the logic after GUI is created
def load_background_logic():
    print("Loading additional logic (e.g., imports)...")
    global pandas, requests, openpyxl, datetime, math, ET, base64
    
    import requests  # Example of delayed import
    from datetime import datetime
    import math
    import xml.etree.ElementTree as ET
    import base64
    print("Additional logic loaded successfully.")

    
# Set up Tkinter GUI
root = tk.Tk()
root.title("Shopify Tool")

# Let row 0 (text area) expand, and column 0+1 share space
root.rowconfigure(0, weight=1)
root.columnconfigure(0, weight=1)
root.columnconfigure(1, weight=1)

# Text area in row 0
text_area = scrolledtext.ScrolledText(root, wrap=tk.WORD)
text_area.grid(row=0, column=0, columnspan=2, sticky="nsew", padx=10, pady=10)

use_gui_log = False
if use_gui_log:
    sys.stdout = RedirectOutput(text_area)

# Buttons in rows below
download_button = tk.Button(root, text="Download", command=start_download, width=20, height=2)
download_button.grid(row=1, column=0, pady=5)

upload_button = tk.Button(root, text="Upload", command=start_upload, width=20, height=2)
upload_button.grid(row=1, column=1, pady=5)

collection_download_button = tk.Button(root, text="Download Collections", command=start_collection_download, width=20, height=2)
collection_download_button.grid(row=2, column=0, pady=5)

collection_upload_button = tk.Button(root, text="Upload Collections", command=start_collection_upload, width=20, height=2)
collection_upload_button.grid(row=2, column=1, pady=5)

file_alt_download_button = tk.Button(root, text="Download Files Alt Texts", command=lambda: start_detached_task("Download Files Alt Texts", download_shopify_files_alt_texts), width=25, height=2)
file_alt_download_button.grid(row=3, column=0, pady=5)

file_alt_upload_button = tk.Button(root, text="Upload Files Alt Texts", command=lambda: start_detached_task("Upload Files Alt Texts", upload_shopify_files_alt_texts), width=25, height=2)
file_alt_upload_button.grid(row=3, column=1, pady=5)

seo_alt_text_button = tk.Button(root, text="Generate SEO Alt Texts (AI)", command=lambda: start_detached_task("Generate SEO Alt Texts (AI)", generate_seo_alt_texts), width=30, height=2)
seo_alt_text_button.grid(row=4, column=0, columnspan=2, pady=5)

# Set window size
root.geometry("500x700")

# Show the window first, then load heavy logic in the background
root.after(100, lambda: start_detached_task("Background setup", load_background_logic))

# Run the application
root.mainloop()
